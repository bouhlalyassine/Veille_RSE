#  .venv/Scripts/Activate.ps1
#  python -m streamlit run app.py
#  git add .    # git commit -m "Màj"   # git push -u origin master


import unicodedata
from datetime import date, timedelta
from html import escape
from io import BytesIO
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

from Settings import (
    MEDIA_SCOUT_SOURCE_CATALOG,
    MEDIA_SCOUT_URLS,
    MEDIA_SCOUT_THEMES,
    MEDIA_SCOUT_THEME_EMOJI,
    MEDIA_SCOUT_VEILLES,
    MEDIA_SCOUT_VEILLE_EMOJI,
    MEDIA_SCOUT_FORCED_SOURCE_THEMES,
    MEDIA_SCOUT_SOURCE_ZONES,
    compute_signal_du_jour,
    current_cache_slot,
    data_media_scout,
    format_last_update,
    get_source_origin,
    load_css,
    media_scrape_timestamp,
    translate_titles_to_french,
)


# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Veille LDA",
    page_icon="📰",
    layout="wide",
    initial_sidebar_state="collapsed",
)

load_css()


# ─── Pre-warm cache (scraping global au demarrage, cache 10h) ─────────────────
# data_media_scout est l'unique source : decoree @st.cache_data(ttl=36000),
# on l'appelle ici pour peupler le cache au demarrage. Tous les appels suivants
# (rendu du tableau de bord Veille) hit le cache en quelques millisecondes.
with st.spinner("Recherche d'actualités en cours sur l'ensemble des sources (2 mises à jour paramétrées : 7h00 et 19h00) — quelques minutes…"):
    _slot = current_cache_slot()
    data_media_scout(MEDIA_SCOUT_URLS, slot=_slot)
    # Fige l'horodatage de cette collecte (meme slot -> meme valeur jusqu'au prochain creneau)
    media_scrape_timestamp(slot=_slot)


# ─── Mois FR ──────────────────────────────────────────────────────────────────
_MOIS_FR = {
    "January": "janvier", "February": "février", "March": "mars",
    "April": "avril", "May": "mai", "June": "juin",
    "July": "juillet", "August": "août", "September": "septembre",
    "October": "octobre", "November": "novembre", "December": "décembre",
}


def _format_date_fr(value):
    if pd.isna(value):
        return "Date non disponible"
    date_str = pd.Timestamp(value).strftime("%d %B %Y")
    for eng, fr in _MOIS_FR.items():
        date_str = date_str.replace(eng, fr)
    return date_str


def _format_date_short(value):
    if pd.isna(value):
        return "—"
    date_str = pd.Timestamp(value).strftime("%d %b %Y")
    en_mois = {"Jan": "jan.", "Feb": "fév.", "Mar": "mars", "Apr": "avr.",
               "May": "mai", "Jun": "juin", "Jul": "juil.", "Aug": "août",
               "Sep": "sept.", "Oct": "oct.", "Nov": "nov.", "Dec": "déc."}
    for en, fr in en_mois.items():
        date_str = date_str.replace(en, fr)
    return date_str


def _veille_tone(veille: str) -> str:
    return {
        "Veille Reglementaire":   "reg",
        "Veille Informative":     "inf",
        "Veille Evenementielle":  "evt",
        "Veille Concurrentielle": "con",
    }.get(veille, "inf")


def _veille_display(veille: str) -> str:
    """Retourne la version accentuée pour l'affichage."""
    return {
        "Veille Reglementaire":   "Veille Réglementaire",
        "Veille Informative":     "Veille Informative",
        "Veille Evenementielle":  "Veille Évènementielle",
        "Veille Concurrentielle": "Veille Concurrentielle",
    }.get(veille, veille)


def _theme_display(theme: str) -> str:
    return {
        "Agrumes, Fruits rouges & Maraichage":           "🍊 Agrumes, Fruits rouges & Tomates cerises",
        "Elevage (Ovins, Bovins, Caprins, Volailles)":   "🐄 Élevage (Ovins, Bovins, Caprins, Volailles & Aquaculture)",
        "Produits laitiers & Epicerie fine":             "🧀 Produits laitiers & Épicerie fine",
        "Environnement, Eau & Energie":                  "🌍 Environnement, Eau & Énergie",
        "ESG, QSE & SST":                                "🏛️ Normes : ESG, QSE & SST",
    }.get(theme, theme)


def _theme_plain(theme: str) -> str:
    """Nom du theme SANS emoji (le filigrane decoratif fournit deja l'emoji)."""
    emoji = MEDIA_SCOUT_THEME_EMOJI.get(theme, "")
    return _theme_display(theme).replace(emoji, "").strip()


# Couleur d'accent par thème -> dégradé de fond + emoji filigrane des boutons.
_THEME_COLORS = {
    "Agrumes, Fruits rouges & Maraichage":           "#E8833A",  # orange agrumes
    "Elevage (Ovins, Bovins, Caprins, Volailles)":   "#9268C2",  # violet
    "Produits laitiers & Epicerie fine":             "#C9A227",  # ambre/or
    "Environnement, Eau & Energie":                  "#4FA268",  # vert
    "ESG, QSE & SST":                                "#5C84C4",  # bleu normes
}


def _zone_label(source_name: str) -> str:
    return {
        "MAROC": "MA",
        "EU": "UE",
        "WORLD": "WW",
    }.get(MEDIA_SCOUT_SOURCE_ZONES.get(source_name, ""), "")


def _zone_priority(source_name: str) -> int:
    """Priorité d'affichage : Maroc (0) > EU (1) > World (2) > inconnu (3)."""
    return {
        "MAROC": 0,
        "EU": 1,
        "WORLD": 2,
    }.get(MEDIA_SCOUT_SOURCE_ZONES.get(source_name, ""), 3)


# ─── CSS — design tokens + components ─────────────────────────────────────────
def _design_css(dark: bool = False, frozen: bool = True) -> str:
    if dark:
        palette = """
:root{
  --paper:#2E2A24;
  --paper-2:#39342D;
  --paper-3:#443E35;
  --ink:#EDE3CC;
  --ink-2:#D2C3A4;
  --ink-3:#9C8C71;
  --ink-4:#776B55;
  --line:#48413A;
  --line-soft:#3E382F;
  --line-strong:#7C6D4E;
  --gold:#C2A66B;
  --gold-deep:#A89060;
  --plum:#B59CCA;
  --plum-soft:#3F3349;
  --alert:#C76858;
  --alert-soft:#4D2922;
  --alert-deep:#E8C9BD;
  --teal:#80A0AE;
  --teal-soft:#2C4654;
  --teal-deep:#A8C2CD;
  --teal-darker:#C5DBE3;
  --green:#95AB75;
  --green-soft:#33401E;
  --green-deep:#DCEAC5;
  --tracking:.14em;
  --tracking-loose:.22em;
}
"""
    else:
        palette = """
:root{
  --paper:#FDFBF6;
  --paper-2:#F7F2E6;
  --paper-3:#EFE6CC;
  --ink:#1A120A;
  --ink-2:#3A2F22;
  --ink-3:#7A6A52;
  --ink-4:#A89878;
  --line:#E7D9B2;
  --line-soft:#F1E6CC;
  --line-strong:#C5A96D;
  --gold:#9B8450;
  --gold-deep:#6B5C38;
  --plum:#3D1E50;
  --plum-soft:#E2D2EA;
  --alert:#B03025;
  --alert-soft:#F4D7CC;
  --alert-deep:#7E1F18;
  --teal:#3B6070;
  --teal-soft:#CFE0E6;
  --teal-deep:#1F3A47;
  --teal-darker:#0A1F28;
  --green:#5E7A3A;
  --green-soft:#D9E5C6;
  --green-deep:#162310;
  --tracking:.14em;
  --tracking-loose:.22em;
}
"""
    return """
<style>
@import url('https://fonts.googleapis.com/css2?family=Cinzel:wght@500;600;700&family=Cormorant+Garamond:ital,wght@0,500;0,600;1,500;1,600&family=Inter:wght@400;500;600;700&display=swap');
""" + palette + """

/* Global body & app frame */
html, body, [data-testid="stAppViewContainer"], [data-testid="stMain"] {
    background: var(--paper) !important;
    color: var(--ink);
    font-family: "Inter", system-ui, sans-serif !important;
}
[data-testid="stHeader"]{ display:none !important; }
[data-testid="stToolbar"]{ display:none !important; }
[data-testid="stAppViewContainer"]{ padding-top:0 !important; top:0 !important; }
[data-testid="stMain"]{ padding-top:0 !important; }
/* Hide default streamlit padding so brand strip can hug top */
div[data-testid="stAppViewContainer"] main .block-container {
    padding-top: 0.5rem !important;
    padding-left: 2.5rem !important;
    padding-right: 2.5rem !important;
    max-width: 1480px !important;
}
[data-testid="stSidebar"] { display: none !important; }
[data-testid="collapsedControl"] { display: none !important; }
.main-pad { max-width: 1380px; margin: 0 auto; }

/* Override main.css legacy rules : left-align + restore normal p weight/size */
div[data-testid="stMarkdownContainer"] {
    text-align: left !important;
    line-height: 1.55 !important;
}
div[data-testid="stMarkdownContainer"] p {
    font-family: "Inter", system-ui, sans-serif !important;
    font-weight: 400 !important;
    font-size: 14px !important;
    line-height: 1.55 !important;
    color: var(--ink-2) !important;
}
div[data-testid="stMarkdownContainer"] ul {
    list-style-type: disc !important;
    margin-left: 1.2em !important;
}
div[data-testid="stMarkdownContainer"] h1,
div[data-testid="stMarkdownContainer"] h2,
div[data-testid="stMarkdownContainer"] h3,
div[data-testid="stMarkdownContainer"] h4 {
    font-family: "Cinzel", serif !important;
    font-weight: 600 !important;
    letter-spacing: .10em !important;
    text-transform: uppercase !important;
    color: var(--ink) !important;
}

/* Round toggle buttons (dark mode + freeze) — uniform style */
.st-key-dark_mode_toggle button,
.st-key-freeze_toggle button{
    background:var(--paper-2) !important;
    border:1px solid var(--line) !important;
    border-radius:50% !important;
    width:38px !important;
    height:38px !important;
    padding:0 !important;
    font-size:16px !important;
    color:var(--ink) !important;
    box-shadow:none !important;
}
.st-key-dark_mode_toggle button:hover,
.st-key-freeze_toggle button:hover{
    background:var(--paper-3) !important;
    border-color:var(--line-strong) !important;
}

/* ────────── FILTER BAR ────────── */
""" + ("""
/* Mode FROZEN : position fixed (bulletproof), gap minimal, offset 3px du top */
.st-key-filter-row{
    position:fixed !important;
    top:5px !important;
    left:2.5rem !important;
    right:2.5rem !important;
    max-width:1400px !important;
    margin:0 auto !important;
    z-index:9999 !important;
    background:var(--paper) !important;
    border:2px solid var(--line-strong) !important;
    border-radius:12px !important;
    padding:8px 16px !important;
    box-shadow:0 4px 22px -10px rgba(74,64,48,.45) !important;
    backdrop-filter:blur(8px);
}
/* Compense la barre fixed + 3px offset : pousse le contenu juste sous */
div[data-testid="stAppViewContainer"] main .block-container{
    padding-top:70px !important;
}
""" if frozen else """
/* Mode LIBERE : 5px du top, espacement natif entre les cadres conserve */
.st-key-filter-row{
    position:static !important;
    margin:0 !important;
    background:var(--paper) !important;
    border:2px solid var(--line-strong) !important;
    border-radius:12px !important;
    padding:10px 16px !important;
    box-shadow:0 2px 14px -10px rgba(74,64,48,.30) !important;
}
/* 5px exact entre l'extremite haute du viewport et la filterbar */
div[data-testid="stAppViewContainer"] main .block-container{
    padding-top:5px !important;
}
/* PAS de gap-killer global : les cadres (Signal, Veille x4) gardent leur espacement naturel comme en pinned */
""") + """
.st-key-filter-row [data-testid="stHorizontalBlock"] { align-items:center; }

/* Bump font dans la filterbar (date, themes, toggle) + tout passe en bold */
.st-key-filter-row,
.st-key-filter-row *{
    font-weight:700 !important;
}
.st-key-filter-row [data-testid="stDateInput"] input,
.st-key-filter-row [data-testid="stDateInput"] [data-baseweb="input"]{
    font-size:14.5px !important;
    font-weight:700 !important;
    color:var(--ink) !important;
    height:40px !important;
}
.st-key-filter-row [data-baseweb="select"]{
    min-height:40px !important;
}
.st-key-filter-row [data-baseweb="select"] *,
.st-key-filter-row [data-baseweb="select"] [data-baseweb="placeholder"],
.st-key-filter-row [data-baseweb="select"] input{
    font-size:14px !important;
    font-weight:700 !important;
    color:var(--ink) !important;
}
/* Filtres "categorie" + "type de produit" (dialog Concurrentielle) : placeholder
   plus sombre, identique a celui de "Choisir un theme" (var(--ink)) */
.st-key-con_ptype_filter [data-baseweb="select"] [data-baseweb="placeholder"],
.st-key-con_ptype_filter [data-baseweb="placeholder"],
.st-key-con_pcat_filter [data-baseweb="select"] [data-baseweb="placeholder"],
.st-key-con_pcat_filter [data-baseweb="placeholder"]{
    color:var(--ink) !important;
    -webkit-text-fill-color:var(--ink) !important;
    opacity:1 !important;
    font-weight:700 !important;
}
.st-key-con_ptype_filter [data-baseweb="select"] input::placeholder,
.st-key-con_pcat_filter [data-baseweb="select"] input::placeholder{
    color:var(--ink) !important;
    -webkit-text-fill-color:var(--ink) !important;
    opacity:1 !important;
}
.st-key-filter-row [data-baseweb="tag"]{
    font-size:13.5px !important;
    font-weight:700 !important;
    padding:3px 12px !important;
}
.st-key-filter-row [data-testid="stToggle"] label p,
.st-key-filter-row [data-testid="stToggle"] label div,
[data-testid="stToggle"] label p{
    font-size:14.5px !important;
    font-weight:700 !important;
    color:var(--ink) !important;
}
.st-key-filter-row [data-testid="stButton"] button{
    font-weight:700 !important;
}

/* Derniere MaJ : caption compacte (label + date/heure), centree H + V */
/* Hook stable via container keye (st-key-lastupd-box) — plus fiable que :has().
   On force tout le chemin DOM (block -> element -> markdown) a pleine largeur,
   hauteur 100%, marges nulles, et on centre H + V a chaque niveau. */
.st-key-lastupd-box{
    display:flex !important;
    flex-direction:column !important;
    justify-content:center !important;   /* centrage vertical */
    align-items:center !important;        /* centrage horizontal */
    height:100% !important;
    min-height:40px !important;
    margin:0 !important;
    gap:0 !important;
}
.st-key-lastupd-box [data-testid="stElementContainer"],
.st-key-lastupd-box [data-testid="stMarkdown"],
.st-key-lastupd-box [data-testid="stMarkdownContainer"]{
    width:100% !important;
    margin:0 !important;
    padding:0 !important;
}
.last-update{
    display:flex !important;
    flex-direction:column;
    justify-content:center;   /* centrage vertical interne */
    align-items:center;       /* centrage horizontal interne */
    text-align:center;
    width:100% !important;
    margin:0 !important;
    line-height:1.25;
    gap:2px;
}
.last-update .lu-label{
    font-size:14px !important;
    letter-spacing:.04em;
    text-transform:uppercase;
    color:var(--ink-3) !important;
    font-weight:700 !important;
    white-space:nowrap;
}
.last-update .lu-value{
    font-size:13.5px !important;
    color:var(--gold-deep) !important;
    font-weight:800 !important;
    white-space:nowrap;
}

/* Calendar popover : meilleur contraste (jours, semaines, header) */
[data-baseweb="calendar"]{
    background:var(--paper) !important;
    color:var(--ink) !important;
}
[data-baseweb="calendar"] *{ color:var(--ink) !important; }
[data-baseweb="calendar"] [aria-label*="week"],
[data-baseweb="calendar"] [role="columnheader"]{
    color:var(--ink-2) !important;
    font-weight:700 !important;
}
/* Dates hors plage (avant min_value, apres max_value, mois adjacents) : grisees */
[data-baseweb="calendar"] [aria-disabled="true"],
[data-baseweb="calendar"] [disabled],
[data-baseweb="calendar"] button[disabled],
[data-baseweb="calendar"] button[aria-disabled="true"],
[data-baseweb="calendar"] [aria-label*="outside"],
[data-baseweb="calendar"] [aria-label*="Outside"]{
    color:var(--ink-4) !important;
    opacity:.30 !important;
    background:var(--line-soft) !important;
    cursor:not-allowed !important;
    text-decoration:line-through !important;
    text-decoration-thickness:1px !important;
    text-decoration-color:var(--ink-4) !important;
}
/* Empeche le hover/click visuel sur dates desactivees */
[data-baseweb="calendar"] [aria-disabled="true"]:hover,
[data-baseweb="calendar"] button[disabled]:hover,
[data-baseweb="calendar"] button[aria-disabled="true"]:hover{
    background:var(--line-soft) !important;
    transform:none !important;
}
[data-baseweb="calendar"] [aria-selected="true"]{
    background:var(--gold) !important;
    color:#FDFBF6 !important;
}
[data-baseweb="popover"]{
    background:var(--paper) !important;
    border:1px solid var(--line-strong) !important;
}

/* Multiselect placeholder + chips : meilleur contraste */
[data-baseweb="select"] [data-baseweb="placeholder"],
[data-baseweb="select"] input::placeholder{
    color:var(--ink-2) !important;
    font-weight:500 !important;
    opacity:1 !important;
}
[data-baseweb="select"] [role="option"]{
    color:var(--ink) !important;
}
[data-baseweb="menu"]{
    background:var(--paper) !important;
    border:1px solid var(--line-strong) !important;
}
[data-baseweb="menu"] [role="option"]:hover{
    background:var(--paper-2) !important;
}

/* Bouton navigation À Propos / Veille DA — 2 etats bien distincts & visibles */
/* Base commune (forme, taille) */
.st-key-nav_apropos button,
.st-key-nav_veille button{
    border-radius:8px !important;
    font-weight:800 !important;
    min-height:40px !important;
    white-space:nowrap !important;
    box-shadow:0 4px 12px -8px rgba(74,64,48,.45) !important;
    transition:filter .15s ease, background .15s ease, border-color .15s ease, transform .15s ease !important;
}
.st-key-nav_apropos button p,
.st-key-nav_veille button p{ font-weight:800 !important; font-size:14px !important; }

/* Les 2 etats (À Propos ET Veille DA) ont le MEME style : contour gold + fond
   ambré clair au repos -> remplissage gold + texte crème au survol. */
.st-key-nav_apropos button,
.st-key-nav_veille button{
    background:var(--paper-3) !important;
    border:1.5px solid var(--gold-deep) !important;
}
.st-key-nav_apropos button p,
.st-key-nav_veille button p{ color:var(--gold-deep) !important; }
.st-key-nav_apropos button:hover,
.st-key-nav_veille button:hover{
    background:var(--gold-deep) !important;
    border-color:var(--gold-deep) !important;
    transform:translateY(-1px) !important;
}
.st-key-nav_apropos button:hover p,
.st-key-nav_veille button:hover p{ color:var(--paper) !important; }
.st-key-filter-row [data-baseweb="select"] {
    background:var(--paper) !important;
    border-color:var(--line) !important;
    border-radius:8px !important;
    min-height:36px !important;
}
.st-key-filter-row [data-baseweb="select"]:hover {
    border-color:var(--line-strong) !important;
}
/* Multiselect tags as chips */
.st-key-filter-row [data-baseweb="tag"] {
    background:var(--paper-2) !important;
    border:1px solid var(--line-strong) !important;
    border-radius:999px !important;
    color:var(--gold-deep) !important;
    font-family:"Inter" !important;
    font-weight:500 !important;
    font-size:12px !important;
    padding:2px 10px !important;
    margin:2px !important;
}
.st-key-filter-row [data-baseweb="tag"] span { color:inherit !important; }
.st-key-filter-row [data-baseweb="tag"] svg { fill:var(--gold-deep) !important; }
/* Date input */
.st-key-filter-row [data-testid="stDateInput"] input,
.st-key-filter-row [data-testid="stDateInput"] [data-baseweb="input"] {
    background:var(--paper) !important;
    border-color:var(--line) !important;
    border-radius:8px !important;
    color:var(--ink-2) !important;
    font-family:"Inter" !important;
    font-size:13px !important;
    font-weight:500 !important;
    height:36px !important;
}

/* ────────── SIGNAL DU JOUR (full width) ────────── */
.signal{
    position:relative;
    border:4px solid var(--alert);
    border-radius:14px;
    background:linear-gradient(135deg, var(--alert-soft) 0%, var(--paper) 60%);
    padding:22px 28px 22px 32px;
    margin-bottom:24px;
    box-shadow:0 1px 0 0 rgba(176,48,37,.05), 0 18px 36px -24px rgba(176,48,37,.35);
    overflow:hidden;
    display:flex; gap:20px; align-items:stretch;
    transition:transform 160ms ease, box-shadow 160ms ease, border-color 160ms ease;
}
.signal:hover{
    transform:translateY(-1px);
    border-color:var(--alert-deep) !important;
    box-shadow:0 2px 0 0 rgba(176,48,37,.07), 0 22px 44px -22px rgba(176,48,37,.45) !important;
}
.signal::before{
    content:""; position:absolute; left:0; top:0; bottom:0; width:5px;
    background:var(--alert);
}
.signal-num{
    font-family:"Cormorant Garamond", serif; font-style:italic; font-weight:600;
    font-size:96px; line-height:.85; color:var(--alert);
    flex:none; align-self:flex-start; padding-top:8px;
}
.signal-body{ min-width:0; flex:1; }
.signal-eyebrow{
    font-family:"Inter" !important; font-weight:800 !important; font-size:13px !important;
    letter-spacing:.22em !important; text-transform:uppercase !important;
    color:var(--alert) !important; display:flex; flex-wrap:wrap; gap:8px; align-items:center;
}
.signal-eyebrow .badge{
    display:inline-block; background:rgba(176,48,37,.14); color:var(--alert);
    padding:4px 12px; border-radius:5px; font-weight:800 !important; font-size:13px !important;
    letter-spacing:.22em;
}
.signal-eyebrow .cat{ color:var(--alert); font-weight:700; font-size:13px; }
.signal-eyebrow .sep{ color:var(--ink-4); margin:0 2px; }
.signal-eyebrow .info-icon{
    display:inline-flex; align-items:center; justify-content:center;
    width:20px; height:20px; border-radius:50%;
    border:1.5px solid var(--alert); color:var(--alert);
    font-family:"Inter"; font-weight:800; font-size:12px;
    cursor:help; line-height:1; background:transparent;
    transition:background 140ms ease;
}
.signal-eyebrow .info-icon:hover{ background:var(--alert); color:var(--paper); }
.signal-source{
    display:inline-flex; align-items:center; gap:6px; margin-top:14px;
    padding:7px 14px; border:1.5px solid var(--alert);
    border-radius:8px; background:transparent; color:var(--alert) !important;
    font-family:"Inter" !important; font-weight:700 !important; font-size:13px !important;
    letter-spacing:.04em; text-decoration:none !important;
    transition:background 140ms ease, color 140ms ease;
}
.signal-source:hover{ background:var(--alert); color:var(--paper) !important; }
.signal h2{
    font-family:"Cinzel", serif !important; font-weight:700 !important; font-size:24px !important;
    letter-spacing:.10em !important; text-transform:uppercase !important;
    margin:10px 0 12px !important; color:var(--ink) !important; line-height:1.28 !important;
}
.signal h2 .accent{ color:var(--alert); }
.signal p{
    margin:0 !important; font-size:15px !important;
    color:var(--ink-2) !important; line-height:1.6 !important;
    font-family:"Inter", system-ui, sans-serif !important;
}

/* ────────── CADRES (4 Veilles) ────────── */
.cadres-grid-marker{ display:none; }

/* All four card containers - equal height */
.st-key-cadre-reg,
.st-key-cadre-inf,
.st-key-cadre-evt,
.st-key-cadre-con {
    position:relative;
    border:4px solid var(--line-strong) !important;
    border-radius:14px !important;
    background:var(--paper) !important;
    box-shadow:0 1px 0 0 rgba(74,64,48,.04), 0 14px 30px -22px rgba(74,64,48,.35) !important;
    padding:0 !important;
    overflow:hidden;
    height:100% !important;
    display:flex !important;
    flex-direction:column !important;
    transition:transform 160ms ease, box-shadow 160ms ease, border-color 160ms ease;
}
/* Streamlit inner wrapper must also stretch */
.st-key-cadre-reg > div,
.st-key-cadre-inf > div,
.st-key-cadre-evt > div,
.st-key-cadre-con > div {
    display:flex !important;
    flex-direction:column !important;
    flex:1 1 auto !important;
    min-height:0 !important;
}
.st-key-cadre-reg [data-testid="stVerticalBlock"],
.st-key-cadre-inf [data-testid="stVerticalBlock"],
.st-key-cadre-evt [data-testid="stVerticalBlock"],
.st-key-cadre-con [data-testid="stVerticalBlock"]{
    flex:1 1 auto !important; display:flex !important; flex-direction:column !important;
}
/* Subtle hover : lift + intensified shadow + soft tonal border */
.st-key-cadre-reg:hover,
.st-key-cadre-inf:hover,
.st-key-cadre-con:hover{
    transform:translateY(-2px) !important;
}
.st-key-cadre-reg:hover{
    border-color:var(--gold) !important;
    box-shadow:0 2px 0 0 rgba(149,108,42,.05), 0 22px 42px -22px rgba(149,108,42,.42) !important;
}
.st-key-cadre-inf:hover{
    border-color:var(--plum) !important;
    box-shadow:0 2px 0 0 rgba(95,55,90,.05), 0 22px 42px -22px rgba(95,55,90,.42) !important;
}
.st-key-cadre-con:hover{
    border-color:var(--green) !important;
    box-shadow:0 2px 0 0 rgba(64,98,58,.05), 0 22px 42px -22px rgba(64,98,58,.42) !important;
}

/* Tone headers - the .cadre-head is rendered as HTML inside each container */
.cadre-head{
    padding:16px 22px 14px;
    display:flex; align-items:center; justify-content:space-between; gap:14px;
    border-bottom:1px solid var(--line-soft);
}
.cadre-head.reg{ background:linear-gradient(180deg, var(--paper-3) 0%, var(--paper) 100%); }
.cadre-head.inf{ background:linear-gradient(180deg, var(--plum-soft) 0%, var(--paper) 100%); }
.cadre-head.evt{ background:linear-gradient(180deg, var(--teal-soft) 0%, var(--paper) 100%); }
.cadre-head.con{ background:linear-gradient(180deg, var(--green-soft) 0%, var(--paper) 100%); }
.cadre-head h3{
    font-family:"Cinzel", serif !important; font-weight:700 !important; font-size:16px !important;
    letter-spacing:.14em !important; text-transform:uppercase !important;
    margin:0 !important; color:var(--ink) !important; line-height:1.2 !important;
}
.cadre-head.reg h3{ color:var(--gold-deep) !important; }
.cadre-head.inf h3{ color:var(--plum) !important; }
.cadre-head.evt h3{ color:var(--teal) !important; }
.cadre-head.con h3{ color:var(--green) !important; }
.cadre-count{ text-align:right; flex:none; }
.cadre-count .n{
    font-family:"Cinzel", serif; font-weight:700; font-size:30px; line-height:1; color:var(--ink);
}
.cadre-count .l{
    font-size:11px; letter-spacing:.16em; text-transform:uppercase;
    color:var(--ink-3); font-weight:700; margin-top:2px;
}

/* Entries list (font bumped: 14.5px) */
.cadre-list{ padding:10px 22px 6px; display:flex; flex-direction:column; gap:2px; }
.cadre-list .entry{
    display:flex; gap:12px; align-items:flex-start; padding:10px 0;
    border-bottom:1px dashed var(--line-soft);
    font-family:"Inter", system-ui, sans-serif;
    font-size:14.5px; line-height:1.55; color:var(--ink-2);
}
.cadre-list .entry:last-child{ border-bottom:none; }
.cadre-list .entry .b{
    flex:none; width:7px; height:7px; border-radius:50%;
    border:1.5px solid var(--gold); background:transparent; margin-top:8px;
}
.cadre-head.inf ~ .cadre-list .entry .b{ border-color:var(--plum); }
.cadre-head.evt ~ .cadre-list .entry .b{ border-color:var(--teal); }
.cadre-head.con ~ .cadre-list .entry .b{ border-color:var(--green); }
.cadre-list .entry .txt{ flex:1; min-width:0; }
.cadre-list .entry .txt b.zone-tag{ font-weight:700; color:var(--ink); }
.cadre-list .entry .src-num{
    display:inline-block; border:1px solid var(--line-strong); background:var(--paper);
    color:var(--gold-deep); border-radius:4px; padding:0 6px;
    font-family:"Inter"; font-weight:600; font-size:11px; text-decoration:none;
    margin-left:4px; line-height:1.4;
}
.cadre-list .entry .src-num:hover{ background:var(--paper-2); }
.cadre-list .empty-state{
    padding:20px 0; color:var(--ink-3); font-style:italic;
    font-family:"Cormorant Garamond", serif; font-size:15px;
}

/* ────────── "Voir tout" link (pill tonale en footer du cadre) ────────── */
.st-key-open_veille_reg,
.st-key-open_veille_inf,
.st-key-open_veille_con{
    text-align:right !important;
    padding:10px 22px 16px !important;
    margin-top:auto !important;
}
.st-key-open_veille_reg [data-testid="stButton"],
.st-key-open_veille_inf [data-testid="stButton"],
.st-key-open_veille_con [data-testid="stButton"]{
    display:flex !important;
    justify-content:flex-end !important;
}
.st-key-open_veille_reg button[kind="tertiary"],
.st-key-open_veille_inf button[kind="tertiary"],
.st-key-open_veille_con button[kind="tertiary"]{
    background:var(--paper) !important;
    border:1.5px solid var(--line-strong) !important;
    border-radius:999px !important;
    box-shadow:0 1px 0 0 rgba(74,64,48,.05), 0 2px 8px -4px rgba(74,64,48,.25) !important;
    padding:7px 16px !important;
    margin:0 !important;
    font-family:"Inter", system-ui, sans-serif !important;
    font-weight:700 !important;
    font-size:12.5px !important;
    letter-spacing:.08em !important;
    text-transform:uppercase !important;
    cursor:pointer !important;
    transition:transform 160ms ease, background 160ms ease, color 160ms ease, border-color 160ms ease, box-shadow 160ms ease !important;
}
/* Tonal palette per-veille : border + color tonal au repos */
.st-key-open_veille_reg button[kind="tertiary"]{
    color:var(--gold-deep) !important;
    border-color:var(--gold) !important;
}
.st-key-open_veille_inf button[kind="tertiary"]{
    color:var(--plum) !important;
    border-color:var(--plum) !important;
}
.st-key-open_veille_con button[kind="tertiary"]{
    color:var(--green) !important;
    border-color:var(--green) !important;
}
/* Hover : bord tonal mis en valeur (epais + shadow tonale), pas de bg fill */
.st-key-open_veille_reg button[kind="tertiary"]:hover,
.st-key-open_veille_inf button[kind="tertiary"]:hover,
.st-key-open_veille_con button[kind="tertiary"]:hover{
    background:var(--paper) !important;
    border-width:2.5px !important;
    transform:translateX(3px) !important;
}
.st-key-open_veille_reg button[kind="tertiary"]:hover{
    color:var(--gold) !important;
    border-color:var(--gold) !important;
    box-shadow:0 0 0 3px rgba(149,108,42,.12), 0 6px 18px -8px rgba(149,108,42,.45) !important;
}
.st-key-open_veille_inf button[kind="tertiary"]:hover{
    color:var(--plum) !important;
    border-color:var(--plum) !important;
    box-shadow:0 0 0 3px rgba(95,55,90,.14), 0 6px 18px -8px rgba(95,55,90,.45) !important;
}
.st-key-open_veille_con button[kind="tertiary"]:hover{
    color:var(--green) !important;
    border-color:var(--green) !important;
    box-shadow:0 0 0 3px rgba(64,98,58,.14), 0 6px 18px -8px rgba(64,98,58,.45) !important;
}
.st-key-open_veille_reg button[kind="tertiary"]:focus,
.st-key-open_veille_inf button[kind="tertiary"]:focus,
.st-key-open_veille_con button[kind="tertiary"]:focus{
    outline:none !important;
}

/* ────────── Generic buttons (date/toggle context) ────────── */
[data-testid="stButton"] > button:not([kind="tertiary"]) {
    border:1px solid var(--line);
    border-radius:8px;
    background:var(--paper);
    color:var(--ink-2);
    font-family:"Inter";
    font-weight:500;
    font-size:12px;
    padding:7px 12px;
    cursor:pointer;
}
[data-testid="stButton"] > button:not([kind="tertiary"]):hover {
    background:var(--paper-2);
    border-color:var(--line-strong);
    color:var(--ink);
}

/* Toggle gold */
[data-testid="stToggle"] [role="switch"][aria-checked="true"]{
    background:var(--gold) !important;
}
[data-testid="stToggle"] [role="switch"]{
    border-color:var(--line-strong) !important;
}
[data-testid="stToggle"] label p {
    font-family:"Inter" !important;
    font-weight:600 !important;
    font-size:12px !important;
    color:var(--ink-2) !important;
    letter-spacing:.02em !important;
}

/* ────────── DIALOG (expanders) ────────── */
[data-testid="stDialog"]{
    background: var(--paper) !important;
}
/* Masque le titre par defaut "Detail de la veille" */
[data-testid="stDialog"] > div > div:first-child > div:first-child{
    display:none !important;
}
[data-testid="stDialog"] h1,
[data-testid="stDialog"] h2,
[data-testid="stDialog"] h3{
    font-family:"Cinzel", serif !important;
    font-weight:700 !important;
    letter-spacing:.10em !important;
    color: var(--ink) !important;
}
[data-testid="stDialog"] .scout-theme-divider{
    color:var(--gold-deep);
    font-family:"Inter";
    font-size:13px;
    font-weight:700;
    letter-spacing:.16em;
    text-transform:uppercase;
    border-bottom:1px solid var(--line);
    margin:14px 0 8px;
    padding-bottom:6px;
}
[data-testid="stDialog"] .scout-theme-divider .theme-count{
    color:var(--ink-4);
    font-size:11px;
    font-weight:600;
    margin-left:0.4rem;
}
[data-testid="stExpander"] {
    border-radius:10px !important;
    border:1px solid var(--line) !important;
    background:var(--paper) !important;
    margin-bottom:6px !important;
}
[data-testid="stExpander"] summary {
    font-family:"Inter" !important;
    font-weight:600 !important;
    font-size:13px !important;
    color:var(--ink-2) !important;
    padding:8px 12px !important;
}
[data-testid="stExpander"] summary:hover {
    background:rgba(155,132,80,.06) !important;
}
.scout-expander-summary {
    color:var(--ink-2); font-size:14px; line-height:1.55;
    margin-bottom:0.55rem; font-family:"Inter";
}
.scout-expander-link {
    color:var(--gold-deep); font-family:"Inter"; font-size:13px;
    font-weight:600; text-decoration:none; letter-spacing:.06em;
}
.scout-expander-link:hover { color:var(--gold); text-decoration:underline; }

/* ────────── PRESENTATION PANELS (toggle off) ────────── */
.presentation-card{
    background:var(--paper);
    border:4px solid var(--line-strong);
    border-left:6px solid var(--gold);
    border-radius:12px;
    padding:22px 26px;
    margin-bottom:0;
    height:100%;
    min-height:230px;
    box-shadow:0 1px 0 0 rgba(74,64,48,.04), 0 8px 18px -14px rgba(74,64,48,.22);
}
.presentation-card .eyebrow{ color:var(--gold-deep); margin-bottom:12px;
    font-family:"Inter" !important; font-weight:800 !important; font-size:15px !important;
    letter-spacing:.18em !important; text-transform:uppercase !important;
}
.presentation-card h4{
    font-family:"Cinzel", serif !important; font-weight:600 !important; font-size:16px !important;
    letter-spacing:.12em !important; text-transform:uppercase !important;
    margin:0 0 12px !important; color:var(--ink) !important; line-height:1.3 !important;
}
.presentation-card p, .presentation-card li{
    font-family:"Inter" !important; font-size:14px !important; line-height:1.6 !important;
    color:var(--ink-2) !important; margin:0 !important;
}
.presentation-card ol, .presentation-card ul{
    margin:6px 0 0 18px !important; padding:0 !important;
}
.presentation-card li{ margin-bottom:6px !important; }
.presentation-card b{ color:var(--ink); font-weight:600; }

/* ────────── LANDING : boutons de thème (dégradé + emoji filigrane) ────────── */
/* Base commune. Le dégradé de fond + l'emoji filigrane (::after) sont injectés
   dynamiquement par thème (couleur dédiée) dans le bloc landing. */
[class*="st-key-theme_btn_"] button{
    position:relative !important;          /* ancre le filigrane ::after */
    overflow:hidden !important;            /* clippe le filigrane */
    border:1.5px solid var(--gold-deep) !important;
    border-radius:16px !important;
    min-height:155px !important;
    padding:16px 18px !important;
    display:flex !important; align-items:center !important; justify-content:center !important;
    white-space:normal !important;
    text-align:center !important;
    box-shadow:0 6px 18px -14px rgba(74,64,48,.35) !important;
    transition:transform .15s ease, box-shadow .15s ease, border-color .15s ease, background .25s ease !important;
}
[class*="st-key-theme_btn_"] button:hover{
    transform:translateY(-4px) !important;
    box-shadow:0 18px 34px -16px rgba(74,64,48,.55) !important;
    border-color:var(--gold-deep) !important;
}
[class*="st-key-theme_btn_"] button p{
    position:relative !important; z-index:1 !important;   /* texte au-dessus du filigrane */
    font-family:"Inter" !important; font-weight:800 !important;
    font-size:21px !important; line-height:1.4 !important;
    color:var(--ink) !important;
    text-shadow:0 1px 3px var(--paper) !important;        /* lisibilité sur le dégradé */
}

/* Sources section heading */
.sources-heading{
    font-family:"Cinzel", serif !important; font-weight:700 !important; font-size:20px !important;
    letter-spacing:.14em !important; text-transform:uppercase !important; color:var(--ink) !important;
    text-align:center; margin:36px 0 14px;
    padding-top:26px; border-top:1px solid var(--line); position:relative;
}
.sources-heading::after{
    content:""; position:absolute; left:50%; transform:translateX(-50%); top:-1px;
    width:120px; height:2px; background:var(--gold);
}
.sources-heading small{
    display:block; font-family:"Cormorant Garamond", serif !important; font-style:italic;
    font-weight:600 !important; font-size:18px !important; letter-spacing:.04em;
    color:var(--ink-2) !important; text-transform:none; margin-top:10px;
}

/* Sources dataframe : harmonise au theme papier + bordure or (light + dark) */
[data-testid="stDataFrame"]{
    border:3px solid var(--line-strong) !important;
    border-radius:12px !important;
    overflow:hidden !important;
    background:var(--paper) !important;
    box-shadow:0 1px 0 0 rgba(74,64,48,.04), 0 10px 22px -16px rgba(74,64,48,.28) !important;
    /* Glide DataGrid CSS variables : pilote les couleurs du canvas interne */
    --gdg-bg-cell:var(--paper) !important;
    --gdg-bg-cell-medium:var(--paper-2) !important;
    --gdg-bg-header:var(--paper-2) !important;
    --gdg-bg-header-has-focus:var(--paper-3) !important;
    --gdg-bg-header-hovered:var(--paper-3) !important;
    --gdg-border-color:var(--line-strong) !important;
    --gdg-horizontal-border-color:var(--line-soft) !important;
    --gdg-text-dark:var(--ink) !important;
    --gdg-text-light:var(--ink) !important;
    --gdg-text-medium:var(--ink-2) !important;
    --gdg-text-header:var(--ink) !important;
    --gdg-text-header-selected:var(--ink) !important;
    --gdg-text-group-header:var(--ink) !important;
    --gdg-accent-color:var(--gold) !important;
    --gdg-accent-light:var(--gold-deep) !important;
    --gdg-bg-bubble:var(--paper-2) !important;
    --gdg-bg-bubble-selected:var(--gold) !important;
    --gdg-link-color:var(--gold-deep) !important;
    --gdg-cell-horizontal-padding:12px !important;
    --gdg-header-bottom-border-color:var(--line-strong) !important;
}
[data-testid="stDataFrame"] [data-testid="stTableStyledTable"],
[data-testid="stDataFrame"] table{
    background:var(--paper) !important;
    color:var(--ink) !important;
}
[data-testid="stDataFrame"] [role="columnheader"],
[data-testid="stDataFrame"] thead th{
    background:var(--paper-2) !important;
    color:var(--ink) !important;
    font-family:"Cinzel", serif !important;
    font-weight:700 !important;
    letter-spacing:.10em !important;
    text-transform:uppercase !important;
    font-size:11.5px !important;
    border-bottom:2px solid var(--line-strong) !important;
    padding:10px 12px !important;
}
[data-testid="stDataFrame"] [role="row"]{
    border-bottom:1px solid var(--line-soft) !important;
}
[data-testid="stDataFrame"] [role="gridcell"],
[data-testid="stDataFrame"] tbody td{
    color:var(--ink-2) !important;
    font-family:"Inter", sans-serif !important;
    font-size:13px !important;
    padding:8px 12px !important;
    background:var(--paper) !important;
}
[data-testid="stDataFrame"] [role="row"]:nth-child(even) [role="gridcell"],
[data-testid="stDataFrame"] tbody tr:nth-child(even) td{
    background:var(--paper-2) !important;
}
[data-testid="stDataFrame"] [role="row"]:hover [role="gridcell"],
[data-testid="stDataFrame"] tbody tr:hover td{
    background:var(--paper-3) !important;
}
[data-testid="stDataFrame"] a{
    color:var(--gold-deep) !important;
    font-weight:600 !important;
    text-decoration:none !important;
}
[data-testid="stDataFrame"] a:hover{
    color:var(--gold) !important;
    text-decoration:underline !important;
}

/* Bouton "Telecharger (Excel)" — meme langage visuel que les cadres */
.st-key-dl_sources_xlsx button{
    background:var(--paper-2) !important;
    border:2px solid var(--line-strong) !important;
    border-radius:10px !important;
    color:var(--gold-deep) !important;
    font-family:"Inter", sans-serif !important;
    font-weight:700 !important;
    font-size:13px !important;
    letter-spacing:.04em !important;
    height:42px !important;
    box-shadow:0 2px 8px -4px rgba(74,64,48,.18) !important;
    transition:background 160ms ease, color 160ms ease, border-color 160ms ease, transform 160ms ease;
}
.st-key-dl_sources_xlsx button:hover{
    background:var(--gold) !important;
    color:#FDFBF6 !important;
    border-color:var(--gold-deep) !important;
    transform:translateY(-1px) !important;
}

/* ────────── Bouton Veille Evenementielle (CERCLE teal centre dans 20% col) ────────── */
.st-key-open_events_btn{
    height:100% !important;
    display:flex !important;
    align-items:center !important;
    justify-content:center !important;
}
.st-key-open_events_btn > div,
.st-key-open_events_btn [data-testid="stButton"]{
    height:100% !important;
    display:flex !important;
    align-items:center !important;
    justify-content:center !important;
    width:100% !important;
}
.st-key-open_events_btn button{
    width:200px !important;
    height:200px !important;
    border-radius:50% !important;
    background:linear-gradient(135deg, var(--teal-soft) 0%, var(--paper) 90%) !important;
    border:4px solid var(--teal) !important;
    color:var(--ink) !important;
    font-family:"Cinzel", serif !important;
    font-weight:700 !important;
    font-size:13px !important;
    letter-spacing:.14em !important;
    text-transform:uppercase !important;
    padding:0 !important;
    white-space:pre-line !important;
    line-height:1.45 !important;
    text-align:center !important;
    display:flex !important;
    flex-direction:column !important;
    align-items:center !important;
    justify-content:center !important;
    box-shadow:0 1px 0 0 rgba(59,96,112,.06), 0 14px 30px -22px rgba(59,96,112,.45) !important;
    transition:transform 160ms ease, box-shadow 160ms ease, background 160ms ease, color 160ms ease, border-color 160ms ease;
}
.st-key-open_events_btn button > div,
.st-key-open_events_btn button [data-testid="stMarkdownContainer"],
.st-key-open_events_btn button p{
    text-align:center !important;
    margin:0 !important;
    width:100% !important;
}
.st-key-open_events_btn button:hover{
    transform:translateY(-2px) scale(1.03) !important;
    box-shadow:0 2px 0 0 rgba(59,96,112,.08), 0 22px 44px -20px rgba(59,96,112,.60) !important;
}
.st-key-open_events_btn button:disabled{
    opacity:.55 !important;
    cursor:not-allowed !important;
}
.st-key-open_events_btn button p,
.st-key-open_events_btn button div{
    white-space:pre-line !important;
    font-family:"Cinzel", serif !important;
}

/* ────────── Bouton Veille Concurrentielle (CERCLE green) ─────────────────── */
.st-key-open_con_btn{
    height:100% !important;
    display:flex !important;
    align-items:center !important;
    justify-content:center !important;
}
.st-key-open_con_btn > div,
.st-key-open_con_btn [data-testid="stButton"]{
    height:100% !important;
    display:flex !important;
    align-items:center !important;
    justify-content:center !important;
    width:100% !important;
}
.st-key-open_con_btn button{
    width:200px !important;
    height:200px !important;
    border-radius:50% !important;
    background:linear-gradient(135deg, var(--green-soft) 0%, var(--paper) 90%) !important;
    border:4px solid var(--green) !important;
    color:var(--ink) !important;
    font-family:"Cinzel", serif !important;
    font-weight:700 !important;
    font-size:13px !important;
    letter-spacing:.14em !important;
    text-transform:uppercase !important;
    padding:0 !important;
    white-space:pre-line !important;
    line-height:1.45 !important;
    text-align:center !important;
    display:flex !important;
    flex-direction:column !important;
    align-items:center !important;
    justify-content:center !important;
    box-shadow:0 1px 0 0 rgba(64,98,58,.06), 0 14px 30px -22px rgba(64,98,58,.45) !important;
    transition:transform 160ms ease, box-shadow 160ms ease, background 160ms ease, color 160ms ease, border-color 160ms ease;
}
.st-key-open_con_btn button > div,
.st-key-open_con_btn button [data-testid="stMarkdownContainer"],
.st-key-open_con_btn button p{
    text-align:center !important;
    margin:0 !important;
    width:100% !important;
}
.st-key-open_con_btn button:hover{
    transform:translateY(-2px) scale(1.03) !important;
    box-shadow:0 2px 0 0 rgba(64,98,58,.08), 0 22px 44px -20px rgba(64,98,58,.60) !important;
}
.st-key-open_con_btn button:disabled{
    opacity:.55 !important;
    cursor:not-allowed !important;
}
.st-key-open_con_btn button p,
.st-key-open_con_btn button div{
    white-space:pre-line !important;
    font-family:"Cinzel", serif !important;
}

/* ────────── Dialog events : 2 cards par ligne ────────── */
[data-testid="stDialog"] .events-list{
    display:grid;
    grid-template-columns:repeat(2, 1fr);
    gap:12px;
    margin-top:8px;
}
[data-testid="stDialog"] .event-card{
    display:flex !important;
    align-items:center;
    gap:14px;
    padding:12px 14px;
    border:2px solid var(--teal);
    border-radius:12px;
    background:var(--teal-soft);
    text-decoration:none !important;
    color:var(--ink) !important;
    transition:transform 160ms ease, border-color 160ms ease, box-shadow 160ms ease, background 160ms ease;
    min-height:108px;
}
[data-testid="stDialog"] .event-card:hover{
    border-color:var(--teal) !important;
    background:var(--paper) !important;
    transform:translateY(-2px);
    box-shadow:0 12px 24px -14px rgba(59,96,112,.55);
}
[data-testid="stDialog"] .event-card .event-date{
    flex:0 0 78px;
    text-align:center;
    border-right:1px solid var(--teal);
    padding-right:10px;
}
[data-testid="stDialog"] .event-card .event-day{
    font-family:"Cinzel", serif;
    font-weight:700;
    font-size:26px;
    line-height:1;
    color:var(--teal);
}
[data-testid="stDialog"] .event-card .event-month{
    font-family:"Inter", sans-serif;
    font-weight:600;
    font-size:10px;
    letter-spacing:.10em;
    text-transform:uppercase;
    color:var(--ink-2);
    margin-top:4px;
}
[data-testid="stDialog"] .event-card .event-body{
    flex:1; min-width:0;
}
[data-testid="stDialog"] .event-card .event-tags{
    display:flex; flex-wrap:wrap; gap:6px; margin-bottom:6px;
}
[data-testid="stDialog"] .event-card .event-cat,
[data-testid="stDialog"] .event-card .event-scope,
[data-testid="stDialog"] .event-card .event-theme{
    display:inline-block;
    font-family:"Inter", sans-serif;
    font-size:9px;
    font-weight:700;
    letter-spacing:.10em;
    text-transform:uppercase;
    padding:2px 7px;
    border-radius:4px;
}
[data-testid="stDialog"] .event-card .event-cat{
    background:var(--paper); color:var(--teal); border:1px solid var(--teal);
}
[data-testid="stDialog"] .event-card .event-scope{
    background:var(--paper); color:var(--gold-deep); border:1px solid var(--gold);
}
[data-testid="stDialog"] .event-card .event-theme{
    background:var(--paper); color:var(--ink-2); border:1px solid var(--line-strong);
}
[data-testid="stDialog"] .event-card h4{
    font-family:"Cormorant Garamond", serif !important;
    font-style:italic;
    font-weight:600 !important;
    font-size:16px !important;
    color:var(--ink) !important;
    margin:3px 0 !important;
    line-height:1.3 !important;
    letter-spacing:0 !important;
    text-transform:none !important;
}
[data-testid="stDialog"] .event-card .event-source{
    font-family:"Inter", sans-serif;
    font-size:11px;
    color:var(--ink-3);
    margin-top:0;
}
[data-testid="stDialog"] .event-card .event-arrow{
    flex:none;
    font-size:22px;
    color:var(--teal);
    font-weight:700;
    padding-left:8px;
}

/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
/* MOBILE RESPONSIVE (tablet & phone, <=768px)                              */
/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
@media (max-width: 768px) {
    /* Mobile : supprimer les espaces (br) entre la topbar et le contenu */
    .st-key-landing-top-spacer,
    .st-key-veille-top-spacer{ display:none !important; }

    /* Collapse les conteneurs INVISIBLES (CSS injecté <style> + iframes JS height=0)
       qui occupent quand même un "slot" flex avec gap -> espaces fantômes entre la
       topbar et le contenu. display:none retire le slot ; les <style> restent actifs
       (métadonnées non affectées par le display) et les iframes JS continuent de
       s'exécuter (toujours présents dans le DOM). */
    [data-testid="stElementContainer"]:has(style),
    [data-testid="stElementContainer"]:has(iframe),
    .element-container:has(style),
    .element-container:has(iframe){
        display:none !important;
    }

    /* Main container : reduire padding lateral pour gagner de l espace */
    /* padding-top reduit car filterbar passe en flow sur mobile (position:relative) */
    .stApp .block-container,
    div[data-testid="stAppViewContainer"] main .block-container{
        padding-left:12px !important;
        padding-right:12px !important;
        padding-top:12px !important;
        max-width:100% !important;
    }
    /* Hero header reduit */
    .hero-title{ font-size:1.85rem !important; line-height:1.15 !important; }
    .hero-subtitle{ font-size:.9rem !important; }
    .hero-eyebrow{ font-size:11px !important; letter-spacing:.18em !important; }

    /* ── Filterbar : passe en mode relatif (non fixed) sur mobile, scroll naturel ── */
    .st-key-filter-row{
        position:relative !important;
        top:auto !important;
        left:auto !important;
        right:auto !important;
        padding:10px 12px !important;
        z-index:auto !important;
        border-radius:10px !important;
    }
    .st-key-filter-row [data-testid="stHorizontalBlock"]{
        flex-wrap:wrap !important;
        gap:8px !important;
    }
    .st-key-filter-row [data-testid="stDateInput"] input,
    .st-key-filter-row [data-testid="stDateInput"] [data-baseweb="input"]{
        font-size:13px !important;
        height:38px !important;
    }
    .st-key-filter-row [data-baseweb="select"]{ min-height:38px !important; }
    .st-key-filter-row [data-baseweb="select"] *,
    .st-key-filter-row [data-baseweb="select"] input{ font-size:13px !important; }
    .st-key-filter-row [data-testid="stToggle"] label p{ font-size:13px !important; }

    /* Cacher le bouton freeze sur mobile (filterbar deja en flow) */
    .st-key-freeze_toggle{ display:none !important; }
    /* Compresser legerement dark mode toggle */
    .st-key-dark_mode_toggle button{
        padding:4px 8px !important;
        min-width:38px !important;
        min-height:38px !important;
    }

    /* ── Signal du jour : compact + retirer le "!" pour gagner de l'espace ── */
    .signal{
        padding:16px 14px !important;
        gap:12px !important;
        flex-direction:row !important;
    }
    .signal-num{
        display:none !important;
    }
    .signal h2{ font-size:17px !important; line-height:1.3 !important; }
    .signal p{ font-size:13.5px !important; line-height:1.5 !important; }
    .signal .badge{ font-size:10px !important; padding:3px 8px !important; letter-spacing:.10em !important; }
    .signal .info-icon{ width:16px !important; height:16px !important; font-size:10px !important; }
    .signal .signal-source{ font-size:12px !important; }
    .signal .signal-eyebrow{ gap:6px !important; flex-wrap:wrap !important; }

    /* ── Cercles → CARRES A BORDS ARRONDIS pleine largeur sur mobile ────── */
    /* Plus de circle 50% : devient rectangle 100% largeur, padding interne genereux */
    .st-key-open_events_btn,
    .st-key-open_con_btn{
        width:100% !important;
        max-width:none !important;
        height:auto !important;
    }
    .st-key-open_events_btn > div,
    .st-key-open_con_btn > div,
    .st-key-open_events_btn [data-testid="stButton"],
    .st-key-open_con_btn [data-testid="stButton"]{
        width:100% !important;
        height:auto !important;
    }
    .st-key-open_events_btn button,
    .st-key-open_con_btn button{
        width:100% !important;
        max-width:none !important;
        min-height:130px !important;
        height:auto !important;
        border-radius:16px !important;
        font-size:13.5px !important;
        line-height:1.45 !important;
        letter-spacing:.10em !important;
        border-width:3px !important;
        padding:24px 28px !important;
        white-space:pre-line !important;
    }
    /* Force centrage du label */
    .st-key-open_events_btn button p,
    .st-key-open_con_btn button p,
    .st-key-open_events_btn button > div,
    .st-key-open_con_btn button > div,
    .st-key-open_events_btn button [data-testid="stMarkdownContainer"],
    .st-key-open_con_btn button [data-testid="stMarkdownContainer"]{
        text-align:center !important;
        margin:0 !important;
        width:100% !important;
    }

    /* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
    /* REORDRE MOBILE : Signal -> Inf -> Reg -> Concurrentielle -> Evt    */
    /* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
    /* Decomposer les stHorizontalBlock contenant nos cadres/cercles via :has */
    /* Cela rend les colonnes filles directement enfants du parent stVerticalBlock */
    [data-testid="stHorizontalBlock"]:has(> div > [data-testid="stColumn"] .st-key-signal-wrap),
    [data-testid="stHorizontalBlock"]:has(> div > [data-testid="stColumn"] .st-key-open_con_btn),
    [data-testid="stHorizontalBlock"]:has(> div > [data-testid="stColumn"] .st-key-open_events_btn),
    [data-testid="stHorizontalBlock"]:has(> div > [data-testid="stColumn"] .st-key-cadre-inf),
    [data-testid="stHorizontalBlock"]:has(> div > [data-testid="stColumn"] .st-key-cadre-reg){
        display:contents !important;
    }
    /* Les colonnes deviennent pleine largeur */
    [data-testid="stColumn"]:has(.st-key-signal-wrap),
    [data-testid="stColumn"]:has(.st-key-open_con_btn),
    [data-testid="stColumn"]:has(.st-key-open_events_btn),
    [data-testid="stColumn"]:has(.st-key-cadre-inf),
    [data-testid="stColumn"]:has(.st-key-cadre-reg){
        width:100% !important;
        max-width:100% !important;
        flex:1 1 100% !important;
        min-width:0 !important;
        padding:0 !important;
        margin-bottom:14px !important;
    }
    /* Ordre d'affichage mobile via CSS order */
    [data-testid="stColumn"]:has(.st-key-signal-wrap){ order:1 !important; }
    [data-testid="stColumn"]:has(.st-key-cadre-inf){ order:2 !important; }
    [data-testid="stColumn"]:has(.st-key-cadre-reg){ order:3 !important; }
    [data-testid="stColumn"]:has(.st-key-open_con_btn){ order:4 !important; }
    [data-testid="stColumn"]:has(.st-key-open_events_btn){ order:5 !important; }

    /* ── Cadres veille (Reg + Inf) : font reduits ── */
    .cadre-head{ padding:12px 14px 10px !important; }
    .cadre-head h3{ font-size:13.5px !important; letter-spacing:.10em !important; }
    .cadre-count .n{ font-size:22px !important; }
    .cadre-count .l{ font-size:9.5px !important; letter-spacing:.10em !important; }
    .cadre-list{ padding:8px 14px !important; }
    .cadre-list .entry{ font-size:13px !important; padding:8px 0 !important; gap:8px !important; }
    .cadre-list .entry .src-num{ font-size:10px !important; padding:0 5px !important; }
    .cadre-list .empty-state{ font-size:13px !important; padding:14px 0 !important; }

    /* ── "Voir tout" pill : plus compact ── */
    .st-key-open_veille_reg button[kind="tertiary"],
    .st-key-open_veille_inf button[kind="tertiary"],
    .st-key-open_veille_con button[kind="tertiary"]{
        font-size:11px !important;
        padding:6px 12px !important;
        letter-spacing:.06em !important;
    }

    /* ── Dialog : full-width sur mobile + events grid 1 colonne ── */
    [data-testid="stDialog"] > div{ max-width:96vw !important; padding:14px !important; }
    [data-testid="stDialog"] .events-list{
        grid-template-columns:1fr !important;
        gap:10px !important;
    }
    [data-testid="stDialog"] .event-card{
        min-height:auto !important;
        padding:10px 12px !important;
        gap:10px !important;
    }
    [data-testid="stDialog"] .event-card h4{ font-size:14px !important; }
    [data-testid="stDialog"] .event-card .event-day{ font-size:22px !important; }
    [data-testid="stDialog"] .event-card .event-cat,
    [data-testid="stDialog"] .event-card .event-scope,
    [data-testid="stDialog"] .event-card .event-theme{ font-size:8.5px !important; padding:2px 6px !important; }
    [data-testid="stDialog"] h1, [data-testid="stDialog"] h2, [data-testid="stDialog"] h3{
        font-size:18px !important; letter-spacing:.06em !important;
    }
    [data-testid="stDialog"] [data-testid="stExpander"] summary{
        font-size:12px !important;
        padding:6px 10px !important;
    }

    /* ── Sources table : scroll horizontal natif Streamlit + heading compact ── */
    .sources-heading{ font-size:17px !important; }
    .sources-heading small{ font-size:11px !important; }

    /* ── Presentation cards : padding reduit ── */
    .presentation-card{
        padding:16px 18px !important;
        min-height:auto !important;
        margin-bottom:12px !important;
    }
    .presentation-card .eyebrow{ font-size:11px !important; }
    .presentation-card p,
    .presentation-card li{ font-size:13.5px !important; line-height:1.55 !important; }

    /* ── Dialog scout summary : font lisible ── */
    .scout-expander-summary{ font-size:13px !important; }
    .scout-theme-divider{ font-size:12px !important; }
}

/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
/* SMALL PHONES (<=480px) - encore plus compact                             */
/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */
@media (max-width: 480px) {
    .hero-title{ font-size:1.55rem !important; }
    .hero-subtitle{ font-size:.82rem !important; }

    /* Signal : pas de "!" sur petits écrans non plus */
    .signal{ flex-direction:column !important; align-items:flex-start !important; }
    .signal-num{ display:none !important; }
    .signal h2{ font-size:16px !important; }
    .signal p{ font-size:13px !important; }

    /* Cercles -> carres : hauteur reduite sur petits ecrans */
    .st-key-open_events_btn button,
    .st-key-open_con_btn button{
        min-height:115px !important;
        height:auto !important;
        font-size:12.5px !important;
        padding:20px 22px !important;
    }

    /* Cadres compactes */
    .cadre-head h3{ font-size:12.5px !important; }
    .cadre-list .entry{ font-size:12.5px !important; padding:7px 0 !important; }
    .cadre-list .entry .src-num{ font-size:9.5px !important; }

    /* Dialog encore plus serre */
    [data-testid="stDialog"] > div{ padding:10px !important; }
}
</style>
"""


# ─── Catalogue evenementiel : journees + salons + congres ─────────────────────
# Chaque entree a : month, day, name, category, theme, scope, source, url
#   category : "Journee mondiale" | "Journee nationale" | "Salon" | "Congres" | "Expo" | "Rencontre"
#   theme    : un des MEDIA_SCOUT_THEMES (filtre par les themes selectionnes)
#   scope    : "International" | "Maroc"
#   source   : nom de la source de reference (UN, FAO, UNESCO, organisateur, etc.)
_EVENTS_CATALOG = [
    # ───── Energie & Environnement (eau / climat / energie) ─────
    {"month": 3, "day": 22, "name": "Journée mondiale de l'eau", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/water-day"},
    {"month": 11, "day": 19, "name": "Journée mondiale des toilettes", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/toilet-day"},
    {"month": 6, "day": 17, "name": "Journée lutte contre désertification et sécheresse", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/desertification-day"},
    {"month": 5, "day": 18, "name": "World Water Forum (variable)", "category": "Congrès", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "World Water Council", "url": "https://www.worldwaterforum.org/"},
    {"month": 1, "day": 21, "name": "Salon SITeau Marrakech", "category": "Salon", "theme": "Environnement, Eau & Energie", "scope": "Maroc", "source": "Organisateur SITeau", "url": "https://www.siteau.ma/"},

    # ───── (suite) Climat / Biodiversite ─────
    {"month": 4, "day": 22, "name": "Journée internationale de la Terre", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Earth Day Network", "url": "https://www.earthday.org/"},
    {"month": 6, "day": 5, "name": "Journée mondiale de l'environnement", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "UNEP", "url": "https://www.worldenvironmentday.global/"},
    {"month": 6, "day": 8, "name": "Journée mondiale de l'océan", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/oceans-day"},
    {"month": 5, "day": 22, "name": "Journée internationale de la biodiversité", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/biological-diversity-day"},
    {"month": 9, "day": 7, "name": "Journée internationale de l'air pur", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/clean-air-day"},
    {"month": 9, "day": 16, "name": "Journée internationale de la couche d'ozone", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/ozone-day"},
    {"month": 12, "day": 5, "name": "Journée mondiale des sols", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "FAO", "url": "https://www.fao.org/world-soil-day"},
    {"month": 10, "day": 13, "name": "Journée internationale réduction des risques de catastrophe", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/disaster-reduction-day"},
    {"month": 5, "day": 19, "name": "ChangeNOW Summit Paris", "category": "Congrès", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "ChangeNOW", "url": "https://www.changenow.world/"},
    {"month": 5, "day": 12, "name": "IFAT Munich (environnement et eau)", "category": "Salon", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Messe München", "url": "https://ifat.de/"},
    {"month": 11, "day": 27, "name": "Pollutec Lyon", "category": "Salon", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "RX France", "url": "https://www.pollutec.com/"},
    {"month": 11, "day": 10, "name": "COP30 - Belém", "category": "Congrès", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "UNFCCC", "url": "https://unfccc.int/"},
    {"month": 6, "day": 5, "name": "Salon Pollutec Maroc Casablanca", "category": "Salon", "theme": "Environnement, Eau & Energie", "scope": "Maroc", "source": "RX France", "url": "https://www.pollutec-maroc.com/"},

    # ───── (suite) Énergie & Transition ─────
    {"month": 10, "day": 22, "name": "Journée internationale des énergies renouvelables", "category": "Journée mondiale", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "IRENA", "url": "https://www.irena.org/"},
    {"month": 5, "day": 26, "name": "World Hydrogen Summit Rotterdam", "category": "Congrès", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Sustainable Energy Council", "url": "https://www.world-hydrogen-summit.com/"},
    {"month": 11, "day": 2, "name": "ADIPEC Abu Dhabi", "category": "Salon", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "dmg events", "url": "https://www.adipec.com/"},
    {"month": 5, "day": 6, "name": "Intersolar Europe Munich", "category": "Salon", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Solar Promotion", "url": "https://www.intersolar.de/"},
    {"month": 10, "day": 1, "name": "Renewable Energy Morocco Casablanca", "category": "Salon", "theme": "Environnement, Eau & Energie", "scope": "Maroc", "source": "MASEN", "url": "https://www.masen.ma/"},
    {"month": 12, "day": 1, "name": "World Future Energy Summit Abu Dhabi", "category": "Congrès", "theme": "Environnement, Eau & Energie", "scope": "International", "source": "Masdar", "url": "https://www.worldfutureenergysummit.com/"},

    # ───── Agrumes, Fruits rouges & Maraichage ─────
    {"month": 4, "day": 21, "name": "SIAM - Salon International de l'Agriculture du Maroc Meknès", "category": "Salon", "theme": "Agrumes, Fruits rouges & Maraichage", "scope": "Maroc", "source": "Comité National du SIAM", "url": "https://www.salon-agriculture.ma/"},
    {"month": 10, "day": 16, "name": "Journée mondiale de l'alimentation", "category": "Journée mondiale", "theme": "Agrumes, Fruits rouges & Maraichage", "scope": "International", "source": "FAO", "url": "https://www.fao.org/world-food-day"},
    {"month": 6, "day": 7, "name": "Journée mondiale de la sécurité sanitaire des aliments", "category": "Journée mondiale", "theme": "Agrumes, Fruits rouges & Maraichage", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/food-safety-day"},
    {"month": 5, "day": 20, "name": "Journée mondiale des abeilles", "category": "Journée mondiale", "theme": "Agrumes, Fruits rouges & Maraichage", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/bee-day"},
    {"month": 2, "day": 21, "name": "Salon International de l'Agriculture Paris", "category": "Salon", "theme": "Agrumes, Fruits rouges & Maraichage", "scope": "International", "source": "CENECA", "url": "https://www.salon-agriculture.com/"},
    {"month": 10, "day": 17, "name": "SIAL Paris (alimentation)", "category": "Salon", "theme": "Agrumes, Fruits rouges & Maraichage", "scope": "International", "source": "Comexposium", "url": "https://www.sialparis.com/"},
    {"month": 10, "day": 4, "name": "Anuga Cologne", "category": "Salon", "theme": "Agrumes, Fruits rouges & Maraichage", "scope": "International", "source": "Koelnmesse", "url": "https://www.anuga.com/"},
    {"month": 2, "day": 11, "name": "Salon Halieutis Agadir (pêche et aquaculture)", "category": "Salon", "theme": "Agrumes, Fruits rouges & Maraichage", "scope": "Maroc", "source": "Département de la Pêche Maritime", "url": "https://www.salonhalieutis.com/"},

    # ───── Élevage ─────
    {"month": 9, "day": 15, "name": "SPACE - Salon élevage Rennes", "category": "Salon", "theme": "Elevage (Ovins, Bovins, Caprins, Volailles)", "scope": "International", "source": "SPACE Organization", "url": "https://www.space.fr/"},
    {"month": 11, "day": 10, "name": "EuroTier Hanovre", "category": "Salon", "theme": "Elevage (Ovins, Bovins, Caprins, Volailles)", "scope": "International", "source": "DLG", "url": "https://www.eurotier.com/"},
    {"month": 10, "day": 6, "name": "Sommet de l'Élevage Clermont-Ferrand", "category": "Salon", "theme": "Elevage (Ovins, Bovins, Caprins, Volailles)", "scope": "International", "source": "GIE Sommet de l'Élevage", "url": "https://www.sommet-elevage.fr/"},
    {"month": 5, "day": 28, "name": "VIV Europe (aviculture)", "category": "Salon", "theme": "Elevage (Ovins, Bovins, Caprins, Volailles)", "scope": "International", "source": "VNU Exhibitions", "url": "https://vivworldwide.com/"},

    # ───── Produits laitiers & Epicerie fine (GMS / retail) ─────
    {"month": 9, "day": 27, "name": "GroceryShop Las Vegas", "category": "Congrès", "theme": "Produits laitiers & Epicerie fine", "scope": "International", "source": "Shoptalk", "url": "https://groceryshop.com/"},
    {"month": 3, "day": 15, "name": "ProWein Düsseldorf", "category": "Salon", "theme": "Produits laitiers & Epicerie fine", "scope": "International", "source": "Messe Düsseldorf", "url": "https://www.prowein.com/"},
    {"month": 5, "day": 25, "name": "PLMA Amsterdam (marques de distributeurs)", "category": "Salon", "theme": "Produits laitiers & Epicerie fine", "scope": "International", "source": "PLMA International", "url": "https://www.plmainternational.com/"},
    {"month": 2, "day": 1, "name": "ISM Cologne (confiserie et snacks)", "category": "Salon", "theme": "Produits laitiers & Epicerie fine", "scope": "International", "source": "Koelnmesse", "url": "https://www.ism-cologne.com/"},
    {"month": 5, "day": 21, "name": "Salon Food Morocco Casablanca", "category": "Salon", "theme": "Produits laitiers & Epicerie fine", "scope": "Maroc", "source": "Food Magazine", "url": "https://www.foodmagazine.ma/"},
    {"month": 6, "day": 18, "name": "Halal Morocco Expo Casablanca", "category": "Expo", "theme": "Produits laitiers & Epicerie fine", "scope": "Maroc", "source": "AFICAM", "url": "https://www.halalexpo.ma/"},

    # ───── ESG, QSE & SST ─────
    {"month": 4, "day": 28, "name": "Journée mondiale sécurité et santé au travail", "category": "Journée mondiale", "theme": "ESG, QSE & SST", "scope": "International", "source": "OIT", "url": "https://www.ilo.org/safework/events/safeday"},
    {"month": 9, "day": 22, "name": "UN Global Compact Leaders Summit", "category": "Congrès", "theme": "ESG, QSE & SST", "scope": "International", "source": "UN Global Compact", "url": "https://www.unglobalcompact.org/take-action/leaders-summit"},
    {"month": 10, "day": 7, "name": "Journée mondiale du travail décent", "category": "Journée mondiale", "theme": "ESG, QSE & SST", "scope": "International", "source": "CSI", "url": "https://www.ituc-csi.org/world-day-for-decent-work"},
    {"month": 3, "day": 8, "name": "Journée internationale des droits des femmes", "category": "Journée mondiale", "theme": "ESG, QSE & SST", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/womens-day"},
    {"month": 11, "day": 25, "name": "Journée élimination violences faites aux femmes", "category": "Journée mondiale", "theme": "ESG, QSE & SST", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/ending-violence-against-women-day"},
    {"month": 12, "day": 10, "name": "Journée des droits de l'homme", "category": "Journée mondiale", "theme": "ESG, QSE & SST", "scope": "International", "source": "Nations Unies", "url": "https://www.un.org/en/observances/human-rights-day"},
    {"month": 6, "day": 12, "name": "Journée mondiale contre le travail des enfants", "category": "Journée mondiale", "theme": "ESG, QSE & SST", "scope": "International", "source": "OIT", "url": "https://www.ilo.org/ipec/Campaignandadvocacy/wdacl"},
    {"month": 4, "day": 7, "name": "Journée mondiale de la santé", "category": "Journée mondiale", "theme": "ESG, QSE & SST", "scope": "International", "source": "OMS", "url": "https://www.who.int/campaigns/world-health-day"},
    {"month": 10, "day": 10, "name": "Journée mondiale de la santé mentale", "category": "Journée mondiale", "theme": "ESG, QSE & SST", "scope": "International", "source": "OMS", "url": "https://www.who.int/campaigns/world-mental-health-day"},
    # Maroc - jours nationaux & cérémonies
    {"month": 1, "day": 11, "name": "Manifeste de l'Indépendance", "category": "Journée nationale", "theme": "ESG, QSE & SST", "scope": "Maroc", "source": "Gouvernement du Maroc", "url": "https://www.maroc.ma/"},
    {"month": 5, "day": 1, "name": "Fête du Travail (Maroc)", "category": "Journée nationale", "theme": "ESG, QSE & SST", "scope": "Maroc", "source": "Gouvernement du Maroc", "url": "https://www.maroc.ma/"},
    {"month": 7, "day": 30, "name": "Fête du Trône", "category": "Journée nationale", "theme": "ESG, QSE & SST", "scope": "Maroc", "source": "Cabinet Royal", "url": "https://www.maroc.ma/"},
    {"month": 8, "day": 14, "name": "Allégeance Oued Ed-Dahab", "category": "Journée nationale", "theme": "ESG, QSE & SST", "scope": "Maroc", "source": "Gouvernement du Maroc", "url": "https://www.maroc.ma/"},
    {"month": 8, "day": 20, "name": "Révolution du Roi et du Peuple", "category": "Journée nationale", "theme": "ESG, QSE & SST", "scope": "Maroc", "source": "Gouvernement du Maroc", "url": "https://www.maroc.ma/"},
    {"month": 11, "day": 6, "name": "Marche Verte", "category": "Journée nationale", "theme": "ESG, QSE & SST", "scope": "Maroc", "source": "Gouvernement du Maroc", "url": "https://www.maroc.ma/"},
    {"month": 11, "day": 18, "name": "Fête de l'Indépendance (Maroc)", "category": "Journée nationale", "theme": "ESG, QSE & SST", "scope": "Maroc", "source": "Gouvernement du Maroc", "url": "https://www.maroc.ma/"},
]


def _events_in_window(start_date, end_date, themes=None):
    """Renvoie tous les evenements du catalogue dont la date tombe dans [start, end],
    optionnellement filtres par themes selectionnes. Resultat trie chronologiquement."""
    events = []
    selected_themes = set(themes) if themes else None
    for year in range(start_date.year, end_date.year + 1):
        for event in _EVENTS_CATALOG:
            if selected_themes is not None and event.get("theme") not in selected_themes:
                continue
            evt_date = date(year, event["month"], event["day"])
            if start_date <= evt_date <= end_date:
                events.append({**event, "date": evt_date})
    return sorted(events, key=lambda item: item["date"])


def _future_events_for_themes(start_date, themes, months=12):
    """Evenements a venir dans les 12 mois a partir de start_date, filtres par themes."""
    end_date = start_date + timedelta(days=30 * months)
    return _events_in_window(start_date, end_date, themes=themes)


# Alias retrocompatibilite
_events_between = _events_in_window
_ESG_EVENT_CATALOG = _EVENTS_CATALOG


# Mapping tone -> variables CSS (s'adapte automatiquement light/dark)
_VEILLE_TONE_VARS = {
    "reg": ("--gold-deep", "--paper-3"),
    "inf": ("--plum", "--plum-soft"),
    "evt": ("--teal", "--teal-soft"),
    "con": ("--green", "--green-soft"),
}


# ─── Dialog : detail des articles d'une Veille ────────────────────────────────
@st.dialog(" ", width="large")  # titre vide -> masque via CSS, on affiche notre propre titre
def _show_veille_details(veille, group, selected_themes, calendar_events=None):
    tone = _veille_tone(veille)
    border_var, soft_var = _VEILLE_TONE_VARS.get(tone, ("--gold-deep", "--paper-3"))
    is_dark = st.session_state.get("dark_mode", False)

    # En dark mode : fond neutre (paper-2) + texte clair (ink) pour lisibilite
    # En light mode : fond tinte (soft tone) + texte tone-colore pour effet design
    closed_bg = "--paper-2" if is_dark else soft_var
    closed_text = "--ink" if is_dark else border_var
    open_body_bg = "--paper" if is_dark else "--paper-2"

    # CSS scope par tone : variables CSS pour adaptation auto light/dark
    st.markdown(
        f"""<style>
[data-testid="stDialog"] [data-testid="stExpander"],
[data-testid="stDialog"] details[data-testid="stExpander"],
[data-testid="stDialog"] [data-testid="stExpander"] > div{{
    border:2px solid var({border_var}) !important;
    background:var({closed_bg}) !important;
    border-radius:10px !important;
    margin-bottom:8px !important;
    box-shadow:0 1px 0 0 rgba(0,0,0,.06), 0 4px 12px -10px var({border_var}) !important;
}}
[data-testid="stDialog"] [data-testid="stExpander"] summary,
[data-testid="stDialog"] details[data-testid="stExpander"] > summary{{
    background:var({closed_bg}) !important;
    color:var({closed_text}) !important;
    font-weight:700 !important;
    border-radius:8px 8px 0 0 !important;
}}
[data-testid="stDialog"] [data-testid="stExpander"] summary:hover,
[data-testid="stDialog"] details[data-testid="stExpander"] > summary:hover{{
    filter:brightness(1.05) !important;
}}
/* Expander OUVERT : bandeau tonal en haut + body neutre */
[data-testid="stDialog"] [data-testid="stExpander"][open],
[data-testid="stDialog"] details[data-testid="stExpander"][open]{{
    background:var({open_body_bg}) !important;
}}
[data-testid="stDialog"] [data-testid="stExpander"][open] > summary,
[data-testid="stDialog"] details[data-testid="stExpander"][open] > summary{{
    background:var({border_var}) !important;
    color:var(--paper) !important;
    border-bottom:2px solid var({border_var}) !important;
    border-radius:8px 8px 0 0 !important;
}}
/* Body de l'expander ouvert : fond papier + texte ink-2 */
[data-testid="stDialog"] [data-testid="stExpander"][open] > div:not(summary),
[data-testid="stDialog"] details[data-testid="stExpander"][open] > div{{
    background:var({open_body_bg}) !important;
    color:var(--ink-2) !important;
    padding-top:0.6rem !important;
}}
[data-testid="stDialog"] [data-testid="stExpander"][open] .scout-expander-summary{{
    color:var(--ink-2) !important;
}}
[data-testid="stDialog"] .scout-theme-divider{{
    color:var({border_var}) !important;
    border-bottom-color:var({border_var}) !important;
}}
</style>""",
        unsafe_allow_html=True,
    )

    emoji = MEDIA_SCOUT_VEILLE_EMOJI.get(veille, "")
    st.markdown(f"### {emoji} {_veille_display(veille)}")

    if calendar_events:
        st.markdown(
            '<div class="scout-theme-divider">📅 Calendrier ESG/RSE & journées thématiques</div>',
            unsafe_allow_html=True,
        )
        for event in calendar_events:
            event_url = event.get("url", "")
            header = f"📅  {_format_date_fr(event['date'])}  |  {event['name']}  |  {event['category']}"
            with st.expander(header):
                st.markdown(f"**Catégorie :** {escape(event['category'])}")
                if event_url:
                    st.markdown(
                        f'<a class="scout-expander-link" href="{escape(event_url, quote=True)}" target="_blank" rel="noopener">→ Site officiel</a>',
                        unsafe_allow_html=True,
                    )

    if group.empty:
        if not calendar_events:
            st.info("Aucun article classé dans cette veille sur la période et le thème sélectionné.")
        return

    # ── Veille Concurrentielle : tag « type de produit » + filtre dédié ──
    is_con = (veille == "Veille Concurrentielle")
    if is_con:
        _con_t = _con_theme(selected_themes)
        group = group.copy()
        group["_ptype"] = group.apply(lambda r: _detect_product_type(r, _con_t), axis=1)
        group["_pcat"] = group.apply(lambda r: _detect_category(r, _con_t), axis=1)

        # Diversité totale (décide quels filtres afficher)
        _all_cats = set(group["_pcat"])
        _all_prods = set(group["_ptype"])

        # Sélections courantes (lues AVANT rendu des widgets) -> filtres LIÉS :
        # les options de chaque filtre dépendent de la sélection de l'autre.
        # On purge d'abord toute sélection devenue invalide (ex: changement de
        # thème — les catégories/types diffèrent d'un profil concurrentiel à
        # l'autre), sinon st.multiselect lèverait « value not in options ».
        _cur_cat = [c for c in (st.session_state.get("con_pcat_filter") or []) if c in _all_cats]
        _cur_prod = [t for t in (st.session_state.get("con_ptype_filter") or []) if t in _all_prods]
        if _cur_cat != (st.session_state.get("con_pcat_filter") or []):
            st.session_state["con_pcat_filter"] = _cur_cat
        if _cur_prod != (st.session_state.get("con_ptype_filter") or []):
            st.session_state["con_ptype_filter"] = _cur_prod
        # Catégories dispo = celles présentes pour le(s) produit(s) sélectionné(s)
        _sub_cat = group[group["_ptype"].isin(_cur_prod)] if _cur_prod else group
        # Produits dispo = ceux présents pour la/les catégorie(s) sélectionnée(s)
        _sub_prod = group[group["_pcat"].isin(_cur_cat)] if _cur_cat else group
        # Union avec la sélection courante -> évite l'erreur "value not in options"
        _cat_set = set(_sub_cat["_pcat"]) | set(_cur_cat)
        _prod_set = set(_sub_prod["_ptype"]) | set(_cur_prod)
        _prof = _CON_PROFILES.get(_con_t) or {}
        _cat_order = _prof.get("categories", [])
        _prod_order = _prof.get("product_order", [])
        _present_cat = [c for c in _cat_order if c in _cat_set]
        _present = [t for t in _prod_order if t in _prod_set]

        # Filtres liés Catégorie + Type de produit. Les deux côte à côte si chacun
        # a >1 valeur ; sinon le filtre unique est centré (même largeur qu'un duo).
        _show_cat = len(_all_cats) > 1
        _show_prod = len(_all_prods) > 1
        _selc, _sel = [], []

        def _f_cat(slot):
            with slot:
                return st.multiselect(
                    label="", label_visibility="collapsed",
                    options=_present_cat,
                    placeholder="🗂️ Filtrer par catégorie",
                    default=[],
                    key="con_pcat_filter",
                )

        def _f_prod(slot):
            with slot:
                return st.multiselect(
                    label="", label_visibility="collapsed",
                    options=_present,
                    placeholder="🏷️ Filtrer par type de produit",
                    default=[],
                    key="con_ptype_filter",
                )

        if _show_cat and _show_prod:
            _, fcat, fprod, _ = st.columns([8, 42, 42, 8])
            _selc = _f_cat(fcat)
            _sel = _f_prod(fprod)
        elif _show_cat or _show_prod:
            # Filtre unique centré (mêmes proportions qu'un filtre du duo)
            _, fone, _ = st.columns([29, 42, 29])
            if _show_cat:
                _selc = _f_cat(fone)
            else:
                _sel = _f_prod(fone)

        if _selc:
            group = group[group["_pcat"].isin(_selc)]
        if _sel:
            group = group[group["_ptype"].isin(_sel)]
        if (_selc or _sel) and group.empty:
            st.info("Aucun article concurrent pour la sélection.")
            return

    for theme in selected_themes:
        theme_articles = group[group["Theme"] == theme]
        if theme_articles.empty:
            continue
        st.markdown(
            f'<div class="scout-theme-divider">{escape(_theme_display(theme))}</div>',
            unsafe_allow_html=True,
        )
        # Tri : Zone Maroc > EU > World, puis Date desc
        theme_sorted = theme_articles.copy()
        theme_sorted["_zone_prio"] = theme_sorted["Website_name"].astype(str).map(_zone_priority)
        theme_sorted = theme_sorted.sort_values(["_zone_prio", "Date"], ascending=[True, False])
        for _, row in theme_sorted.iterrows():
            date_str = _format_date_fr(row.get("Date"))
            title = str(row.get("Title", "")).strip()
            source = str(row.get("Website_name", "")).strip()
            desc = str(row.get("Description", "")).strip()
            link = str(row.get("Link", "")).strip()

            # Date + source en gras via markdown (st.expander supporte markdown depuis 1.26+)
            # Pour la Veille Concurrentielle : tags « catégorie » + « type de produit ».
            if is_con:
                _cat = str(row.get("_pcat", "")).strip()
                _typ = str(row.get("_ptype", "")).strip()
                _tags = "  |  ".join(f"`{t}`" for t in (_cat, _typ) if t)
                _tagpart = f"{_tags}  |  " if _tags else ""
                header = f"📰  **{date_str}**  |  {_tagpart}{title}  |  **{source}**"
            else:
                header = f"📰  **{date_str}**  |  {title}  |  **{source}**"
            with st.expander(header):
                if desc:
                    st.markdown(f'<div class="scout-expander-summary">{escape(desc)}</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div class="scout-expander-summary"><i>Pas de résumé disponible pour cet article.</i></div>', unsafe_allow_html=True)
                if link:
                    st.markdown(
                        f'<a class="scout-expander-link" href="{escape(link, quote=True)}" target="_blank" rel="noopener">→ Lire l\'article original</a>',
                        unsafe_allow_html=True,
                    )


# ─── Dialog : Veille Evenementielle (calendrier 12 mois) ──────────────────────
@st.dialog(" ", width="large")
def _show_events_dialog(events, themes, start_date):
    st.markdown(f"### 📅 Veille Évènementielle")
    end_d = start_date + timedelta(days=365)
    st.caption(
        f"{len(events)} événement(s) à venir sur 12 mois "
        f"({_format_date_fr(start_date)} → {_format_date_fr(end_d)}) "
        f"· filtré sur le thème sélectionné."
    )

    if not events:
        st.info("Aucun événement à venir sur le thème sélectionné dans les 12 prochains mois.")
        return

    # Regroupement par mois pour une lecture chronologique
    cards = []
    for evt in events:
        d = evt["date"]
        cat = evt.get("category", "")
        scope = evt.get("scope", "")
        scope_chip = "🇲🇦 Maroc" if scope == "Maroc" else "🌍 International"
        url = evt.get("url", "") or "#"
        source = evt.get("source", "")
        date_fr = _format_date_fr(d)

        card = (
            f'<a class="event-card" href="{escape(url, quote=True)}" target="_blank" rel="noopener">'
            f'<div class="event-date">'
            f'<div class="event-day">{d.day:02d}</div>'
            f'<div class="event-month">{date_fr.split(" ", 1)[1] if " " in date_fr else date_fr}</div>'
            f'</div>'
            f'<div class="event-body">'
            f'<div class="event-tags">'
            f'<span class="event-cat">{escape(cat)}</span>'
            f'<span class="event-scope">{scope_chip}</span>'
            f'</div>'
            f'<h4>{escape(evt["name"])}</h4>'
            f'<div class="event-source">{escape(source)}</div>'
            f'</div>'
            f'<div class="event-arrow">→</div>'
            f'</a>'
        )
        cards.append(card)

    st.markdown(
        '<div class="events-list">' + "".join(cards) + "</div>",
        unsafe_allow_html=True,
    )


def _cadre_synthesis_html(rows, veille_key: str, themes: list) -> str:
    """Liste compacte des titres du TOP 5 d'une veille (tri zone Maroc > UE >
    Monde, puis date), au format « ZONE : Titre » SANS la source.

    Les titres sont garantis en français (traduits via cache si la source est en
    anglais). Remplace l'ancienne synthèse LLM, source de phrases incomplètes /
    incohérentes (« WW : d'énergie propre. (14 words) », etc.).
    """
    if rows.empty:
        return (
            '<div class="cadre-list"><div class="empty-state">'
            'Aucun signal sur cette veille pour la période et le thème sélectionnés.'
            '</div></div>'
        )

    # Tri Maroc > UE > Monde puis date desc, on garde les 5 premiers titres.
    rows_sorted = rows.copy()
    rows_sorted["_zone_prio"] = rows_sorted["Website_name"].astype(str).map(_zone_priority)
    rows_sorted = rows_sorted.sort_values(["_zone_prio", "Date"], ascending=[True, False]).head(5)
    top_rows = list(rows_sorted.iterrows())

    # Traduction FR des titres (cache) -> francais garanti meme pour sources EN.
    titles_fr = translate_titles_to_french(
        tuple(str(r.get("Title", "")).strip() for _, r in top_rows)
    )
    items = []
    for (_, r), title_fr in zip(top_rows, titles_fr):
        title = escape(str(title_fr).strip())
        if len(title) > 180:
            title = title[:177] + "…"
        zone = _zone_label(str(r.get("Website_name", "")))
        zone_prefix = f"<b>{zone}</b> : " if zone else ""
        items.append(
            f'<div class="entry"><span class="b"></span>'
            f'<span class="txt">{zone_prefix}{title}</span></div>'
        )
    return '<div class="cadre-list cadre-list-synthesis">' + "".join(items) + '</div>'


def _cadre_head_html(tone: str, title: str, count: int) -> str:
    return (
        f'<div class="cadre-head {tone}">'
        '<div>'
        f'<h3>{escape(title)}</h3>'
        '</div>'
        '<div class="cadre-count">'
        f'<div class="n">{count}</div>'
        '<div class="l">Articles</div>'
        '</div>'
        '</div>'
    )


# ─── Fragment : isole les rerun des clics "Voir tout" / cercle Evenementielle ──
# Sans @st.fragment, un clic sur un bouton declenche un rerun complet du script
# (re-rendu de tout le layout + filterbar + sources). Avec fragment, seule cette
# portion ré-exécute → suppression du freeze percu apres clic.
def _fold_txt(text: str) -> str:
    """Minuscule + sans accents + apostrophes normalisées (pour matching robuste)."""
    t = unicodedata.normalize("NFKD", str(text).lower())
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t.replace("’", "'").replace("`", "'")


# ════════════════════════════════════════════════════════════════════════════
#  VEILLE CONCURRENTIELLE — profils par thème
#  Chaque thème éligible définit son profil : catégories d'intelligence, mapping
#  source -> catégorie, marques concurrentes (gate de pertinence), sites
#  corporates, marques LDA suivies, et règles de type de produit (tag + filtre).
#  Un article apparaît en Concurrentielle s'il provient d'un site concurrent, OU
#  mentionne un concurrent / une marque LDA, OU porte sur un produit identifiable.
# ════════════════════════════════════════════════════════════════════════════

# ─────────────────────────── T3 : Produits laitiers & Épicerie fine ──────────
_CAT_MARCHE_LOCAL  = "MARCHÉ LOCAL"
_CAT_INTL          = "INTERNATIONAL"
_CAT_FMCG          = "FMCG / RETAIL"
_CAT_NUTRITION     = "NUTRITION / SANTÉ"
_CAT_NOUVEAUTES    = "NOUVEAUTÉS"

_DAIRY_COMPETITORS = (
    # Produits laitiers
    "danone", "copag", "jaouda", "safilait", "jibal", "groupe bel",
    "vache qui rit", "lavachequirit", "kiri", "babybel", "president fromage",
    "lactel", "lactalis", "galbani", "savencia", "caprice des dieux",
    "saint albray", "elle & vire", "nestle maroc", "nido",
    # Distributeurs MDD (laitier + epicerie)
    "label vie", "labelvie", "marjane", "carrefour selection",
    "carrefour maroc", "bim maroc",
    # Epicerie fine
    "aicha", "aïcha", "lesieur cristal", "lesieur ", "zouitina",
    "diana holding", "caracterre", "cartier saada", "sovena",
    "oliveira da serra", "puget huile", "puget olive", "bonne maman",
    "andros", "st dalfour", "st. dalfour", "hero group", "hero confiture",
    "beldimarket",
)
_DAIRY_BRAND_SITES = {"Aïcha", "Nestlé MENA", "Groupe Bel", "Ribambel (Bel)"}
_DAIRY_OWN_BRANDS = ("chergui", "domaines agricoles", "les domaines agricoles", "jaouda")
_DAIRY_SOURCE_CATEGORY = {
    # C1 — Marché local & Maghreb
    "GNews — Presse éco MA": _CAT_MARCHE_LOCAL, "La Vie Éco": _CAT_MARCHE_LOCAL,
    "Le Matin": _CAT_MARCHE_LOCAL, "EcoActu": _CAT_MARCHE_LOCAL,
    "Aujourd'hui Maroc": _CAT_MARCHE_LOCAL,
    "GNews — Centrale Danone": _CAT_MARCHE_LOCAL, "GNews — COPAG Jaouda": _CAT_MARCHE_LOCAL,
    "GNews — Lesieur Cristal": _CAT_MARCHE_LOCAL,
    "GNews — Marjane Maroc": _CAT_MARCHE_LOCAL, "GNews — Olive MA": _CAT_MARCHE_LOCAL,
    "Aïcha": _CAT_MARCHE_LOCAL,
    # C2 — Secteur laitier international
    "GNews — Lait International": _CAT_INTL, "DairyReporter": _CAT_INTL,
    "Food Navigator": _CAT_INTL, "Financial Afrik": _CAT_INTL,
    "GNews — Lactalis": _CAT_INTL, "GNews — Savencia": _CAT_INTL,
    "GNews — Bel Maroc": _CAT_INTL, "Groupe Bel": _CAT_INTL, "Ribambel (Bel)": _CAT_INTL,
    "Nestlé MENA": _CAT_INTL, "GNews — Bonne Maman Andros": _CAT_INTL,
    "GNews — Hero St Dalfour": _CAT_INTL, "GNews — Sovena Puget": _CAT_INTL,
    # C3 — FMCG / Retail / Distribution
    "GNews — FMCG Retail": _CAT_FMCG,
    # C4 — Nutrition fonctionnelle & Santé
    "GNews — Nutrition Santé": _CAT_NUTRITION,
    # C5 — Nouveautés produits & Premium
    "GNews — Nouveautés Premium": _CAT_NOUVEAUTES,
}
_DAIRY_PRODUCT_RULES = [
    ("FROMAGE",     ["fromage", "fromagerie", "cheese", "vache qui rit", "kiri", "babybel",
                     "caprice des dieux", "saint albray", "raclette", "camembert", "emmental",
                     "cancoillotte", "feta", "mozzarella", "cheddar", "gouda"]),
    ("YAOURT",      ["yaourt", "yogurt", "yogourt", "activia", "skyr", "danette",
                     "creme dessert", "lait fermente", "petit suisse"]),
    ("BEURRE/CREME",["beurre", "creme fraiche", "creme legere", " creme ", "butter", " cream"]),
    ("LAIT",        ["lait uht", "lait en poudre", "lait infantile", "lait demi", "lait ecreme",
                     "lait entier", "lait cru", "lait pasteurise", " lait ", "milk", "nido",
                     "poudre de lait", "collecte de lait", "filiere laitiere"]),
    ("HUILE",       ["huile d'olive", "huile de tournesol", "huile de table", "huile vegetale",
                     "huile alimentaire", "huile", "olive oil", "oleicole", "lesieur",
                     "trituration"]),
    ("CONFITURE",   ["confiture", "marmelade", "gelee", "pate a tartiner", "tartiner", "jam"]),
    ("CONSERVE",    ["conserve", "concentre de tomate", "double concentre", "triple concentre",
                     "tomate", "sardine", "thon", "anchois", "harissa", "cornichon",
                     "olive", "olives", "cartier saada"]),
    ("BOISSON",     ["jus de fruit", "jus d'orange", " jus ", "boisson", "nectar", "smoothie", "soda"]),
    ("MIEL",        ["miel", "honey"]),
    ("EPICES/CONDIMENTS", ["epice", "condiment", "sauce", "vinaigre", "moutarde", "mayonnaise", "ketchup"]),
    ("CEREALES",    ["cereale", "farine", "couscous", "semoule", "pates alimentaires", " riz ", "biscuit"]),
]

# ─────────────────── T1 : Agrumes, Fruits rouges & Tomates cerises ───────────
_ACAT_MARCHE      = "MARCHÉ MAROC"
_ACAT_EXPORT      = "EXPORT & MARCHÉS"
_ACAT_CONCURRENTS = "CONCURRENTS"
_ACAT_FILIERE     = "FILIÈRE & PRODUCTION"
_ACAT_PREMIUM     = "VARIÉTÉS & PREMIUM"

_AGRUMES_COMPETITORS = (
    # Exportateurs concurrents (primeurs / agrumes / fruits rouges)
    "azura", "delassus", "duroc", "maraissa", "disma international", "disma",
    "zalar", "rosaflor", "agrumar", "surexport", "driscoll", "soprofel",
    # Organismes filière / export — actus sectorielles pertinentes
    "morocco foodex", "maroc foodex", "eacce", "aspam", "maroc citrus", "apefel",
)
_AGRUMES_OWN_BRANDS = ("les domaines agricoles", "domaines agricoles")
_AGRUMES_SOURCE_CATEGORY = {
    # A1 — Marché Maroc
    "GNews — Agrumes Export MA": _ACAT_MARCHE, "EcoActu": _ACAT_MARCHE,
    "Aujourd'hui Maroc": _ACAT_MARCHE,
    # A2 — Export & marchés internationaux
    "GNews — Marché Agrumes Intl": _ACAT_EXPORT, "GNews — Marché Fruits Rouges": _ACAT_EXPORT,
    "FreshPlaza FR": _ACAT_EXPORT, "Agro-media": _ACAT_EXPORT, "FruitNet": _ACAT_EXPORT,
    "FreshFruitPortal": _ACAT_EXPORT, "Financial Afrik": _ACAT_EXPORT,
    # A3 — Concurrents
    "GNews — Concurrents Primeurs": _ACAT_CONCURRENTS,
    # A4 — Filière & production
    "GNews — Production Fruits MA": _ACAT_FILIERE,
    # A5 — Variétés & premium
    "GNews — Innovations Fruits": _ACAT_PREMIUM,
}
_AGRUMES_PRODUCT_RULES = [
    ("AGRUMES",       ["agrume", "agrumes", "orange", "oranges", "mandarine", "clementine",
                       "citron", "pamplemousse", "pomelo", "lime", "kumquat", "bergamote",
                       "citrus", "navel", "valencia", "nadorcott", "afourer", "maroc late",
                       "soft citrus"]),
    ("TOMATE CERISE", ["tomate cerise", "tomates cerises", "cherry tomato", "cherry tomatoes",
                       "tomate cocktail", "tomate grappe"]),
    ("FRAISE",        ["fraise", "fraises", "strawberry"]),
    ("FRAMBOISE",     ["framboise", "framboises", "raspberry"]),
    ("MYRTILLE",      ["myrtille", "myrtilles", "blueberry"]),
    ("FRUITS ROUGES", ["fruits rouges", "fruit rouge", "mure", "mures", "cassis", "groseille",
                       "groseilles", "blackberry", "cranberry", "berries", "berry",
                       "petits fruits", "soft fruit"]),
    ("TOMATE",        ["tomate", "tomates", "tomato"]),
]

# Registre des profils (clé = thème interne).
_CON_PROFILES = {
    "Produits laitiers & Epicerie fine": {
        "categories": [_CAT_MARCHE_LOCAL, _CAT_INTL, _CAT_FMCG, _CAT_NUTRITION, _CAT_NOUVEAUTES],
        "source_category": _DAIRY_SOURCE_CATEGORY,
        "competitors": _DAIRY_COMPETITORS,
        "brand_sites": _DAIRY_BRAND_SITES,
        "own_brands": _DAIRY_OWN_BRANDS,
        "product_rules": _DAIRY_PRODUCT_RULES,
    },
    "Agrumes, Fruits rouges & Maraichage": {
        "categories": [_ACAT_MARCHE, _ACAT_EXPORT, _ACAT_CONCURRENTS, _ACAT_FILIERE, _ACAT_PREMIUM],
        "source_category": _AGRUMES_SOURCE_CATEGORY,
        "competitors": _AGRUMES_COMPETITORS,
        "brand_sites": set(),
        "own_brands": _AGRUMES_OWN_BRANDS,
        "product_rules": _AGRUMES_PRODUCT_RULES,
    },
}

# Pré-calcul au chargement : ordre des types + mots-clés foldés (1× par profil).
for _prof in _CON_PROFILES.values():
    _prof["product_order"] = [tag for tag, _ in _prof["product_rules"]] + ["AUTRE"]
    _prof["product_rules_folded"] = [
        (tag, [_fold_txt(kw) for kw in kws]) for tag, kws in _prof["product_rules"]
    ]


def _con_theme(selected_themes):
    """Thème concurrentiel actif (le thème sélectionné s'il a un profil), sinon None."""
    if selected_themes:
        t = selected_themes[0]
        if t in _CON_PROFILES:
            return t
    return None


def _detect_category(row, theme) -> str:
    """Catégorie d'intelligence concurrentielle d'un article (selon sa source)."""
    prof = _CON_PROFILES.get(theme)
    if not prof:
        return ""
    return prof["source_category"].get(
        str(row.get("Website_name", "")).strip(), prof["categories"][0]
    )


def _detect_product_type(row, theme) -> str:
    """Retourne un tag de type de produit (selon le profil du thème) ou 'AUTRE'."""
    prof = _CON_PROFILES.get(theme)
    if not prof:
        return "AUTRE"
    text = " " + _fold_txt(str(row.get("Title", "")) + " " + str(row.get("Description", ""))) + " "
    for tag, keywords in prof["product_rules_folded"]:
        for kw in keywords:
            if kw in text:
                return tag
    return "AUTRE"


def _article_mentions_competitor(row, theme) -> bool:
    """Garde l'article seulement s'il est PERTINENT pour les activités LDA du
    thème — sinon écarte le bruit des sources d'intelligence générales.

    Pertinent si :
      1. provient d'un site corporate concurrent, OU
      2. mentionne une marque concurrente / une marque LDA suivie, OU
      3. porte sur un produit identifiable du thème (type ≠ AUTRE).
    """
    prof = _CON_PROFILES.get(theme)
    if not prof:
        return False
    source = str(row.get("Website_name", "")).strip()
    if source in prof["brand_sites"]:
        return True
    text = (str(row.get("Title", "")) + " " + str(row.get("Description", ""))).lower()
    if any(c in text for c in prof["competitors"]):
        return True
    if any(b in text for b in prof["own_brands"]):
        return True
    if _detect_product_type(row, theme) != "AUTRE":
        return True
    return False


@st.fragment
def _render_veille_dashboard(filtered_df, selected_themes, upcoming_events, start_date, signal_html_body):
    # ── ROW 1 : Cercle Concurrentielle | Signal du jour (centre) | Cercle Evenementielle ──
    selected_veille = st.session_state.get("scout_selected_veille")
    n_events = len(upcoming_events)

    # Veille Concurrentielle : active pour les themes dotes d'un profil concurrentiel
    # (Produits laitiers & Epicerie fine ; Agrumes, Fruits rouges & Tomates cerises) ET
    # requiert une mention concurrent / marque LDA / produit identifiable dans l'article.
    con_theme = _con_theme(selected_themes)
    is_con_theme = con_theme is not None
    if is_con_theme:
        con_candidates = filtered_df[filtered_df["Veille"] == "Veille Concurrentielle"]
        # Filtre strict : on ne garde que les articles pertinents pour le thème
        if not con_candidates.empty:
            mask = con_candidates.apply(lambda r: _article_mentions_competitor(r, con_theme), axis=1)
            con_group = con_candidates[mask]
        else:
            con_group = con_candidates
        n_con = len(con_group)
    else:
        con_group = filtered_df.iloc[0:0]  # DataFrame vide avec meme structure
        n_con = 0

    # Layout conditionnel :
    #   - Theme avec profil concurrentiel -> 3 colonnes [Concurrentielle | Signal | Evenementielle]
    #   - Autre theme                     -> 2 colonnes [Signal (etendu) | Evenementielle]
    if is_con_theme:
        sig_cols = st.columns([1.2, 4, 1.2], gap="medium", vertical_alignment="top")
        con_slot = sig_cols[0]
        signal_slot = sig_cols[1]
        evt_slot = sig_cols[2]
    else:
        sig_cols = st.columns([4, 1.2], gap="medium", vertical_alignment="top")
        con_slot = None
        signal_slot = sig_cols[0]
        evt_slot = sig_cols[1]

    if con_slot is not None:
        with con_slot:
            if n_con:
                con_label = f"Veille\nConcurrentielle\n\n{n_con} article{'s' if n_con > 1 else ''}"
                con_help = f"Ouvrir le détail Veille Concurrentielle ({n_con} article(s))"
            else:
                con_label = "Veille\nConcurrentielle\n\nAucun article"
                con_help = "Aucun article concurrent détecté sur la période"
            if st.button(
                con_label,
                key="open_con_btn",
                disabled=(n_con == 0),
                help=con_help,
            ):
                st.session_state["scout_selected_veille"] = "Veille Concurrentielle"
                selected_veille = "Veille Concurrentielle"

    with signal_slot:
        # Container avec key pour pouvoir le cibler en CSS (reordre mobile)
        with st.container(key="signal-wrap"):
            st.markdown(
                '<div class="signal">'
                '<div class="signal-num">!</div>'
                '<div class="signal-body">'
                f'{signal_html_body}'
                '</div>'
                '</div>',
                unsafe_allow_html=True,
            )

    with evt_slot:
        evt_label = (
            f"Veille\nÉvènementielle\n\n{n_events} à venir"
            if n_events else
            "Veille\nÉvènementielle\n\nAucun à venir"
        )
        if st.button(
            evt_label,
            key="open_events_btn",
            disabled=(n_events == 0),
            help="Voir le calendrier des événements à venir sur 12 mois",
        ):
            st.session_state["scout_show_events"] = True

    # ── ROW 2 : Veille Informative (gauche, 50%) + Veille Reglementaire (droite, 50%) ─
    bottom_cols = st.columns(2, gap="medium")
    for idx, (veille_key, slot) in enumerate(zip(
        ["Veille Informative", "Veille Reglementaire"],
        bottom_cols,
    )):
        tone = _veille_tone(veille_key)
        group = filtered_df[filtered_df["Veille"] == veille_key]
        article_count = len(group)
        with slot:
            with st.container(border=True, key=f"cadre-{tone}"):
                st.markdown(
                    _cadre_head_html(tone, _veille_display(veille_key), article_count),
                    unsafe_allow_html=True,
                )
                # Synthese KPI/chiffres/idees generales via LLM (cache par contenu)
                summary_html = _cadre_synthesis_html(group, veille_key, selected_themes)
                st.markdown(summary_html, unsafe_allow_html=True)
                if article_count > 0:
                    if st.button(
                        f"Voir tout ({article_count}) →",
                        key=f"open_veille_{tone}",
                        type="tertiary",
                    ):
                        st.session_state["scout_selected_veille"] = veille_key
                        selected_veille = veille_key

    # ── Dialog handler : detail d'une veille (Reg / Inf / Con) ──
    if selected_veille in ("Veille Reglementaire", "Veille Informative", "Veille Concurrentielle"):
        selected_group = filtered_df[filtered_df["Veille"] == selected_veille]
        # Pour Concurrentielle : filtre strict de pertinence selon le profil du thème
        if selected_veille == "Veille Concurrentielle" and not selected_group.empty and con_theme:
            mask = selected_group.apply(lambda r: _article_mentions_competitor(r, con_theme), axis=1)
            selected_group = selected_group[mask]
        _show_veille_details(selected_veille, selected_group, selected_themes, None)
        st.session_state.pop("scout_selected_veille", None)

    # ── Dialog Evenementielle ──
    if st.session_state.get("scout_show_events"):
        _show_events_dialog(upcoming_events, selected_themes, start_date)
        st.session_state.pop("scout_show_events", None)

    # ── JS Mobile : reordre DOM + force styles via components.html ──
    # IMPORTANT : st.html n'execute PAS les <script> (innerHTML restriction HTML5).
    # components.html cree une iframe avec srcdoc -> les scripts s'executent VRAIMENT.
    # Sur Streamlit Cloud, l'iframe est same-origin (meme app), donc on peut acceder
    # a window.parent.document pour modifier le DOM principal.
    components.html(
        """
<script>
(function(){
  // Helper : applique style avec !important pour battre les CSS rules !important
  function setImp(el, prop, val){
    if (!el) return;
    el.style.setProperty(prop, val, 'important');
  }
  function applyMobileLayout(){
    try {
      var parentWin = window.parent || window;
      var parentDoc = (window.parent && window.parent.document) || document;
      var w = parentWin.innerWidth || parentDoc.documentElement.clientWidth || 1024;
      if (w > 768) return;

      function findCol(key){
        var el = parentDoc.querySelector('.' + key);
        return el ? el.closest('[data-testid="stColumn"]') : null;
      }

      var signalCol = findCol('st-key-signal-wrap');
      var infCol    = findCol('st-key-cadre-inf');
      var regCol    = findCol('st-key-cadre-reg');
      var conCol    = findCol('st-key-open_con_btn');   // optionnel (T3 only)
      var evtCol    = findCol('st-key-open_events_btn');

      // conCol est OPTIONNEL : absent quand le theme n'est pas T3 laitier.
      // Les 4 autres restent obligatoires (signal + inf + reg + evenementiel).
      if (!signalCol || !infCol || !regCol || !evtCol) return;

      // 1. CREER UN CONTAINER EXTERNE hors de l'arbre React Streamlit
      //    React ne suivra pas ce div -> notre reordre survit aux reruns Streamlit
      var stack = parentDoc.getElementById('veille-mobile-stack');
      // Ordre mobile : Signal -> Inf -> Reg -> Concurrentielle (si present) -> Evt
      var ordered = conCol
        ? [signalCol, infCol, regCol, conCol, evtCol]
        : [signalCol, infCol, regCol, evtCol];

      if (!stack) {
        stack = parentDoc.createElement('div');
        stack.id = 'veille-mobile-stack';
        // Inserer juste avant le 1er stHorizontalBlock qui contient une de nos cols
        var anchor = signalCol.closest('[data-testid="stHorizontalBlock"]');
        if (anchor && anchor.parentNode) {
          anchor.parentNode.insertBefore(stack, anchor);
        } else {
          // Fallback : append au body si on ne trouve pas l'ancre
          parentDoc.body.appendChild(stack);
        }
      }
      // Toujours forcer styles flex column 100% sur le stack
      setImp(stack, 'display', 'flex');
      setImp(stack, 'flex-direction', 'column');
      setImp(stack, 'gap', '14px');
      setImp(stack, 'width', '100%');
      setImp(stack, 'max-width', '100%');
      setImp(stack, 'align-items', 'stretch');
      setImp(stack, 'margin', '0');
      setImp(stack, 'padding', '0');

      // 2. Deplacer les 5 cols dans le stack (dans l'ordre voulu)
      //    Verifier si necessaire pour eviter appendChild dans une boucle observer
      var needsReorder = false;
      for (var i = 0; i < ordered.length; i++) {
        if (ordered[i].parentElement !== stack) { needsReorder = true; break; }
      }
      if (needsReorder) {
        ordered.forEach(function(col){ stack.appendChild(col); });
      }

      // 3. Ordre CSS inline (redondance defensive)
      ordered.forEach(function(col, idx){
        setImp(col, 'order', String(idx + 1));
      });

      // 4. Masquer les stHorizontalBlock devenus vides (row1 et row2 d'origine)
      var allHB = parentDoc.querySelectorAll('[data-testid="stHorizontalBlock"]');
      allHB.forEach(function(hb){
        if (hb.querySelectorAll('[data-testid="stColumn"]').length === 0) {
          setImp(hb, 'display', 'none');
        }
      });

      // 5. Largeur 100% sur tous les ancetres du stack jusqu au body
      var ancestor = stack.parentElement;
      var depth = 0;
      while (ancestor && depth < 8 && ancestor.tagName !== 'BODY'){
        setImp(ancestor, 'width', '100%');
        setImp(ancestor, 'max-width', '100%');
        ancestor = ancestor.parentElement;
        depth++;
      }

      // 3. Chaque colonne : pleine largeur + flex pour centrer le contenu
      //    (filtre les eventuels null comme conCol absent)
      [signalCol, infCol, regCol, conCol, evtCol].filter(Boolean).forEach(function(c){
        setImp(c, 'width', '100%');
        setImp(c, 'flex', '1 1 100%');
        setImp(c, 'max-width', '100%');
        setImp(c, 'min-width', '0');
        setImp(c, 'padding', '0');
        setImp(c, 'align-self', 'stretch');
        setImp(c, 'display', 'flex');
        setImp(c, 'flex-direction', 'column');
        setImp(c, 'justify-content', 'center');
        setImp(c, 'align-items', 'center');
      });

      // 4. Cercles -> carres arrondis PLEINE LARGEUR CENTRES
      var btnHeight = w <= 480 ? '95px' : '110px';
      var btnFont = w <= 480 ? '12.5px' : '13.5px';
      ['st-key-open_events_btn', 'st-key-open_con_btn'].forEach(function(key){
        var wrap = parentDoc.querySelector('.' + key);
        if (!wrap) return;
        // Container wrapper : pleine largeur + centre
        setImp(wrap, 'width', '100%');
        setImp(wrap, 'max-width', '100%');
        setImp(wrap, 'display', 'flex');
        setImp(wrap, 'justify-content', 'center');
        setImp(wrap, 'align-items', 'center');
        setImp(wrap, 'margin-left', 'auto');
        setImp(wrap, 'margin-right', 'auto');
        // Inner divs : pleine largeur centres
        Array.prototype.forEach.call(wrap.children, function(child){
          setImp(child, 'width', '100%');
          setImp(child, 'max-width', '100%');
          setImp(child, 'display', 'flex');
          setImp(child, 'justify-content', 'center');
          setImp(child, 'align-items', 'center');
        });
        // stButton wrapper
        var stBtn = wrap.querySelector('[data-testid="stButton"]');
        if (stBtn) {
          setImp(stBtn, 'width', '100%');
          setImp(stBtn, 'max-width', '100%');
          setImp(stBtn, 'display', 'flex');
          setImp(stBtn, 'justify-content', 'center');
        }
        // <button> element : pleine largeur, centre, carre arrondi
        var btn = wrap.querySelector('button');
        if (!btn) return;
        setImp(btn, 'width', '100%');
        setImp(btn, 'max-width', '100%');
        setImp(btn, 'min-width', '0');
        // Hauteur auto (min-height) + padding interne plus genereux
        setImp(btn, 'min-height', w <= 480 ? '115px' : '130px');
        setImp(btn, 'height', 'auto');
        setImp(btn, 'border-radius', '16px');
        setImp(btn, 'padding', w <= 480 ? '20px 22px' : '24px 28px');
        setImp(btn, 'font-size', btnFont);
        setImp(btn, 'margin-left', 'auto');
        setImp(btn, 'margin-right', 'auto');
        setImp(btn, 'display', 'flex');
        setImp(btn, 'align-items', 'center');
        setImp(btn, 'justify-content', 'center');
      });
    } catch(err){
      console.warn('[Mobile layout] error:', err);
    }
  }
  applyMobileLayout();
  setTimeout(applyMobileLayout, 200);
  setTimeout(applyMobileLayout, 600);
  setTimeout(applyMobileLayout, 1500);
  setTimeout(applyMobileLayout, 3000);
  try { (window.parent || window).addEventListener('resize', applyMobileLayout); } catch(e){}

  // MutationObserver dans le PARENT DOC : reapplique des que Streamlit/React modifie
  // (rerun, dialog open/close, etc.). Throttle via requestAnimationFrame.
  try {
    var pDoc = (window.parent && window.parent.document) || document;
    var pWin = window.parent || window;
    if (typeof MutationObserver !== 'undefined' && pWin && !pWin._veilleMobileObs){
      var pending = false;
      pWin._veilleMobileObs = new MutationObserver(function(){
        if (pending) return;
        pending = true;
        (pWin.requestAnimationFrame || window.requestAnimationFrame)(function(){
          pending = false;
          applyMobileLayout();
        });
      });
      pWin._veilleMobileObs.observe(pDoc.body, {
        childList: true,
        subtree: true
      });
    }
  } catch(e){
    console.warn('[Mobile observer] error:', e);
  }
})();
</script>
        """,
        height=0,
    )


# ─── UI STATE (dark mode + filterbar freeze + vue active) ────────────────────
if "dark_mode" not in st.session_state:
    st.session_state["dark_mode"] = False
if "filterbar_frozen" not in st.session_state:
    st.session_state["filterbar_frozen"] = True
# Vue active : "veille" (tableau de bord, par defaut) ou "apropos" (page info agent)
if "scout_view" not in st.session_state:
    st.session_state["scout_view"] = "veille"


def _toggle_dark_mode():
    st.session_state["dark_mode"] = not st.session_state.get("dark_mode", False)


def _toggle_freeze():
    st.session_state["filterbar_frozen"] = not st.session_state.get("filterbar_frozen", True)


def _toggle_view():
    # Bascule entre le tableau de bord Veille et la page À Propos
    cur = st.session_state.get("scout_view", "veille")
    st.session_state["scout_view"] = "apropos" if cur == "veille" else "veille"


def _select_theme(theme: str):
    # Clic sur un bouton de theme (landing) -> pre-selectionne le selectbox
    # et garantit l'affichage du tableau de bord Veille.
    st.session_state["theme_select"] = theme
    st.session_state["scout_view"] = "veille"


# ─── Inject design CSS (palette + filterbar mode) ─────────────────────────────
st.markdown(
    _design_css(
        dark=st.session_state.get("dark_mode", False),
        frozen=st.session_state.get("filterbar_frozen", True),
    ),
    unsafe_allow_html=True,
)


# ─── FILTERBAR (période | thèmes | MàJ | À Propos/Veille DA | dark | freeze) ─
with st.container(key="filter-row"):
    fb_cols = st.columns([1.6, 5.0, 1.5, 1.2, 0.5, 0.5], gap="small", vertical_alignment="center")

    with fb_cols[0]:
        today = date.today()
        days_ago = today - timedelta(days=5)
        date_select = st.date_input(
            label="Période",
            value=(days_ago, today),
            max_value=today,
            min_value=today - timedelta(days=30),
            label_visibility="collapsed",
            format="DD/MM/YYYY",
            help="Période (max 30 jours, défaut 5 jours)",
        )

    with fb_cols[1]:
        # key="theme_select" -> permet aux boutons de theme du landing de pre-selectionner
        # une valeur via st.session_state (callback _select_theme).
        selected_theme = st.selectbox(
            label="Thème",
            options=MEDIA_SCOUT_THEMES,
            index=None,
            key="theme_select",
            format_func=_theme_display,  # _theme_display inclut deja l'emoji
            label_visibility="collapsed",
            placeholder="Choisir un thème",
        )
        # Compat retro : le reste du code attend une liste, on en encapsule la valeur unique
        selected_themes = [selected_theme] if selected_theme else []

    with fb_cols[2]:
        # Dernière mise à jour automatique des données (créneaux 07h / 19h, heure Maroc)
        _last_upd = format_last_update(media_scrape_timestamp(slot=current_cache_slot()))
        with st.container(key="lastupd-box"):
            st.markdown(
                '<div class="last-update" '
                'title="Dernière mise à jour automatique des données — actualisation 2x/jour (07h et 19h, heure Maroc)">'
                f'<span class="lu-label">MàJ 🔄 {_last_upd}</span>'
                #f'<span class="lu-value">{_last_upd}</span>'
                '</div>',
                unsafe_allow_html=True,
            )

    with fb_cols[3]:
        # Bouton de navigation : libelle = la vue VERS laquelle on bascule.
        # Cle distincte par etat -> styles CSS differents (contour vs rempli).
        # Vue "veille" (defaut) -> bouton "À Propos" | Vue "apropos" -> "Veille DA"
        _cur_view = st.session_state.get("scout_view", "veille")
        if _cur_view == "veille":
            st.button(
                "À Propos",
                key="nav_apropos",
                help="En savoir plus sur l'agent de veille",
                on_click=_toggle_view,
                width="stretch",
            )
        else:
            st.button(
                "Veille DA",
                key="nav_veille",
                help="Revenir au tableau de bord de veille",
                on_click=_toggle_view,
                width="stretch",
            )

    with fb_cols[4]:
        _is_dark = st.session_state.get("dark_mode", False)
        st.button(
            "☀️" if _is_dark else "🌙",
            key="dark_mode_toggle",
            help="Basculer mode sombre / clair",
            on_click=_toggle_dark_mode,
            width="stretch",
        )

    with fb_cols[5]:
        _is_frozen = st.session_state.get("filterbar_frozen", True)
        st.button(
            "📌" if _is_frozen else "📍",
            key="freeze_toggle",
            help=("Libérer la barre — elle redeviendra défilante" if _is_frozen
                  else "Figer la barre en haut de page"),
            on_click=_toggle_freeze,
            width="stretch",
        )

# ─── JS Bridge : griser les dates hors plage dans le calendrier ──────────────
# BaseWeb/Streamlit ne grise pas visuellement les dates hors [min_value, max_value].
# Ce script detecte l'ouverture du calendrier via MutationObserver, parse les
# aria-label de chaque bouton-jour, et grise ceux hors plage (futurs / >30j).
components.html(
    f"""
<script>
(function(){{
  function greyOutOfRangeDates(){{
    try {{
      var parentDoc = (window.parent && window.parent.document) || document;
      var calendars = parentDoc.querySelectorAll('[data-baseweb="calendar"]');
      if (!calendars.length) return;
      var today = new Date();
      today.setHours(23, 59, 59, 999);  // inclus aujourd'hui
      var minDate = new Date(today);
      minDate.setDate(minDate.getDate() - 30);
      minDate.setHours(0, 0, 0, 0);
      var monthMap = {{
        'January':0,'February':1,'March':2,'April':3,'May':4,'June':5,
        'July':6,'August':7,'September':8,'October':9,'November':10,'December':11,
        'janvier':0,'fevrier':1,'mars':2,'avril':3,'mai':4,'juin':5,
        'juillet':6,'aout':7,'septembre':8,'octobre':9,'novembre':10,'decembre':11
      }};
      calendars.forEach(function(cal){{
        var buttons = cal.querySelectorAll('button');
        buttons.forEach(function(btn){{
          var label = btn.getAttribute('aria-label') || '';
          // Format attendu (anglais) : "Choose Tuesday, May 13, 2026" ou "May 13, 2026"
          // Format francais : "Mardi 13 mai 2026" ou "13 mai 2026"
          var m = label.match(/(?:[A-Za-z\\u00C0-\\u017F]+,?\\s+)?(\\d+)\\s+(\\w+)\\s+(\\d+)/);
          if (!m) {{
            // Essai format anglais : "May 13, 2026"
            m = label.match(/(\\w+)\\s+(\\d+),?\\s+(\\d+)/);
            if (!m) return;
            var enMonth = m[1].toLowerCase();
            var enDay = parseInt(m[2], 10);
            var enYear = parseInt(m[3], 10);
            var enKey = m[1];  // garde la casse originale pour matcher monthMap
            if (monthMap[enKey] === undefined) {{
              // Essai avec casse normale
              var found = false;
              for (var k in monthMap){{
                if (k.toLowerCase() === enMonth) {{ enKey = k; found = true; break; }}
              }}
              if (!found) return;
            }}
            var dateEn = new Date(enYear, monthMap[enKey], enDay);
            applyGrey(btn, dateEn);
            return;
          }}
          // Format francais : "13 mai 2026"
          var frDay = parseInt(m[1], 10);
          var frMonth = m[2].toLowerCase();
          var frYear = parseInt(m[3], 10);
          var monthIdx = -1;
          for (var k2 in monthMap){{
            if (k2.toLowerCase() === frMonth) {{ monthIdx = monthMap[k2]; break; }}
          }}
          if (monthIdx === -1) return;
          var dateFr = new Date(frYear, monthIdx, frDay);
          applyGrey(btn, dateFr);
        }});
      }});

      function applyGrey(btn, dt){{
        if (dt < minDate || dt > today) {{
          btn.style.setProperty('opacity', '0.30', 'important');
          btn.style.setProperty('background', 'var(--line-soft)', 'important');
          btn.style.setProperty('color', 'var(--ink-4)', 'important');
          btn.style.setProperty('cursor', 'not-allowed', 'important');
          btn.style.setProperty('text-decoration', 'line-through', 'important');
          btn.style.setProperty('pointer-events', 'none', 'important');
        }}
      }}
    }} catch(err){{
      console.warn('[Calendar greyout]', err);
    }}
  }}

  // Run immediately + sur changes du DOM (calendrier ouvert/ferme)
  greyOutOfRangeDates();
  setTimeout(greyOutOfRangeDates, 200);
  setTimeout(greyOutOfRangeDates, 500);

  try {{
    var pDoc = (window.parent && window.parent.document) || document;
    var pWin = window.parent || window;
    if (typeof MutationObserver !== 'undefined' && !pWin._calendarGreyObs){{
      var pending = false;
      pWin._calendarGreyObs = new MutationObserver(function(){{
        if (pending) return;
        pending = true;
        (pWin.requestAnimationFrame || window.requestAnimationFrame)(function(){{
          pending = false;
          greyOutOfRangeDates();
        }});
      }});
      pWin._calendarGreyObs.observe(pDoc.body, {{
        childList: true,
        subtree: true
      }});
    }}
  }} catch(e){{}}
}})();
</script>
    """,
    height=0,
)


# ─── CONTENT ──────────────────────────────────────────────────────────────────
# Vue par defaut = "veille" (tableau de bord actif des l'ouverture, sans toggle).
# La page "apropos" s'affiche uniquement si l'utilisateur clique sur « À Propos ».
if st.session_state.get("scout_view", "veille") == "veille":
    # Espace haut (masqué en mobile via .st-key-veille-top-spacer)
    with st.container(key="veille-top-spacer"):
        st.markdown("<br>", unsafe_allow_html=True)

    invalid_period = (
        not isinstance(date_select, tuple)
        or len(date_select) != 2
        or any(d is None or d == "" for d in date_select)
    )
    issues = []
    if not selected_themes:
        issues.append("un thème")
    if invalid_period:
        issues.append("une période valide (date début + date fin)")

    if issues:
        # Landing visuel : grands boutons (1 par thème) avec dégradé de couleur
        # dédié + emoji en filigrane. Au clic, le thème est pré-sélectionné et le
        # tableau de bord s'affiche. Disposition : 2 centrés en haut, 3 en bas.
        # Espace haut (br) masqué en mobile via .st-key-landing-top-spacer (CSS).
        with st.container(key="landing-top-spacer"):
            st.markdown("<br>", unsafe_allow_html=True)

        # CSS dynamique par thème : dégradé de fond (repos + hover) + filigrane emoji
        _btn_visual_css = "<style>"
        for _bi, _bt in enumerate(MEDIA_SCOUT_THEMES):
            _c = _THEME_COLORS.get(_bt, "#A89060")
            _em = MEDIA_SCOUT_THEME_EMOJI.get(_bt, "")
            _btn_visual_css += (
                f".st-key-theme_btn_{_bi} button{{"
                f"background:linear-gradient(135deg, {_c}3A 0%, {_c}16 48%, var(--paper-3) 100%) !important;}}"
                f".st-key-theme_btn_{_bi} button:hover{{"
                f"background:linear-gradient(135deg, {_c}66 0%, {_c}2E 52%, var(--paper-3) 100%) !important;}}"
                f".st-key-theme_btn_{_bi} button::after{{"
                f"content:'{_em}'; position:absolute; right:-6px; bottom:-26px;"
                f"font-size:104px; line-height:1; opacity:.18; z-index:0;"
                f"pointer-events:none; transform:rotate(-6deg);}}"
                f".st-key-theme_btn_{_bi} button:hover::after{{opacity:.30;}}"
            )
        _btn_visual_css += "</style>"
        st.markdown(_btn_visual_css, unsafe_allow_html=True)

        def _render_theme_btn(theme, idx):
            st.button(
                _theme_plain(theme),  # nom sans emoji (le filigrane fournit l'emoji)
                key=f"theme_btn_{idx}",
                on_click=_select_theme,
                args=(theme,),
                width="stretch",
            )

        _themes = MEDIA_SCOUT_THEMES
        # Ligne 1 : 2 boutons centrés (largeur identique à la ligne du bas via spacers)
        _row1 = st.columns([1, 2, 2, 1], gap="medium")
        _slots1 = [_row1[1], _row1[2]]
        for _j in range(min(2, len(_themes))):
            with _slots1[_j]:
                _render_theme_btn(_themes[_j], _j)
        # Ligne 2 : les 3 restants
        _rest = _themes[2:]
        if _rest:
            _row2 = st.columns(3, gap="medium")
            for _k, _theme in enumerate(_rest[:3]):
                with _row2[_k]:
                    _render_theme_btn(_theme, 2 + _k)
    else:
        with st.spinner("Synthèse des signaux..."):
            media_data_df = data_media_scout(MEDIA_SCOUT_URLS, slot=current_cache_slot())
            start_date, end_date = date_select
            filtered_df = media_data_df[
                (media_data_df["Date"] >= pd.Timestamp(start_date))
                & (media_data_df["Date"] <= pd.Timestamp(end_date))
                & (media_data_df["Theme"].isin(selected_themes))
            ]

        # Veille Evenementielle : evenements a venir 6 mois a partir de start_date,
        # filtres par les themes selectionnes (catalogue statique journees + salons + congres)
        upcoming_events = _future_events_for_themes(start_date, selected_themes, months=12)

        if filtered_df.empty and not upcoming_events:
            st.info("Aucune actualité pertinente trouvée sur la période et les thèmes sélectionnés.")
        else:
            # ── SIGNAL DU JOUR ────────────────────────────────────────────────
            # Pool de candidats. Pour les thèmes à profil concurrentiel (T1/T3),
            # on garantit la présence de la Veille Concurrentielle (pertinente,
            # filtrée) aux côtés du Réglementaire / Informatif — sinon elle est
            # écrasée par le volume des autres veilles et ne pèse jamais.
            _sig_con_theme = _con_theme(selected_themes)
            if _sig_con_theme:
                _pool_parts = []
                for _v in ("Veille Reglementaire", "Veille Informative", "Veille Concurrentielle"):
                    _vsub = filtered_df[filtered_df["Veille"] == _v]
                    if _v == "Veille Concurrentielle" and not _vsub.empty:
                        _vsub = _vsub[_vsub.apply(
                            lambda r: _article_mentions_competitor(r, _sig_con_theme), axis=1
                        )]
                    _pool_parts.append(_vsub.sort_values("Date", ascending=False).head(8))
                _signal_pool = pd.concat(_pool_parts)
                if _signal_pool.empty:
                    _signal_pool = filtered_df.head(25)
            else:
                _signal_pool = filtered_df.head(25)
            signal_articles = tuple(
                " | ".join(
                    part for part in [
                        f"Titre: {str(row.get('Title', '')).strip()}",
                        f"Source: {str(row.get('Website_name', '')).strip()}",
                        f"Date: {_format_date_fr(row.get('Date'))}",
                        f"Theme: {_theme_display(str(row.get('Theme', '')).strip())}",
                        f"Veille: {_veille_display(str(row.get('Veille', '')).strip())}",
                        f"Resume: {str(row.get('Description', '')).strip()}",
                        f"Lien: {str(row.get('Link', '')).strip()}",
                    ]
                    if part and not part.endswith(": ")
                )
                for _, row in _signal_pool.iterrows()
            )
            with st.spinner("Analyse des actualités et identification du signal du jour…"):
                signal = compute_signal_du_jour(signal_articles)

            _signal_tooltip = (
                "Le Signal du jour est identifié automatiquement par l'IA parmi tous les articles "
                "captés sur la période sélectionnée. L'IA évalue chaque article au regard du contexte "
                "et des activités de Les Domaines Agricoles : agriculture, élevage, agro-industrie et "
                "exports. Critères de sélection : nouvelle réglementation UE/Maroc, mise à jour de "
                "normes (ISO, GlobalGAP, IFS, BRC, etc...), décision tarifaire ou accord commercial, "
                "mouvement concurrentiel, évolution des exigences clients export, audit QSE — tout "
                "signal à fort impact potentiel pour les opérations du groupe."
            )

            if signal.get("available") and signal.get("headline"):
                eyebrow = escape(signal.get("eyebrow", ""))
                headline = escape(signal.get("headline", ""))
                body = escape(signal.get("body", ""))
                source_url = signal.get("source_url", "").strip()
                cat_block = f'<span class="sep">·</span><span class="cat">{eyebrow}</span>' if eyebrow else ""
                src_block = (
                    f'<a class="signal-source" href="{escape(source_url, quote=True)}" '
                    f'target="_blank" rel="noopener" title="Ouvrir l\'article d\'origine">'
                    f'🔗 Article d\'origine →</a>'
                    if source_url else ""
                )
                signal_html_body = (
                    '<div class="signal-eyebrow">'
                    '<span class="badge">★ Signal du jour</span>'
                    f'<span class="info-icon" title="{escape(_signal_tooltip, quote=True)}">?</span>'
                    f'{cat_block}'
                    '</div>'
                    f'<h2>{headline}</h2>'
                    f'<p>{body}</p>'
                    f'{src_block}'
                )
            else:
                # Placeholder : cadre toujours visible meme si l'IA n'a rien retourne
                fallback_msg = (
                    "Aucun signal majeur n'a pu être identifié sur la période et les thèmes sélectionnés."
                    if not filtered_df.empty
                    else "Élargissez la période ou ajoutez davantage de thèmes pour faire émerger un signal du jour."
                )
                signal_html_body = (
                    '<div class="signal-eyebrow">'
                    '<span class="badge">★ Signal du jour</span>'
                    f'<span class="info-icon" title="{escape(_signal_tooltip, quote=True)}">?</span>'
                    '</div>'
                    '<h2>Aucun signal critique détecté</h2>'
                    f'<p>{escape(fallback_msg)}</p>'
                )

            # ── Fragment : Signal + Cercle Evt + 3 cadres + dialogs ──
            # Un clic sur "Voir tout" ou le cercle ne ré-execute QUE ce fragment,
            # pas le filterbar/sources/data fetch -> click instantane sans freeze.
            _render_veille_dashboard(
                filtered_df=filtered_df,
                selected_themes=selected_themes,
                upcoming_events=upcoming_events,
                start_date=start_date,
                signal_html_body=signal_html_body,
            )

else:
    # ── Page À PROPOS : présentation de l'agent + sources de référence ───────
    st.markdown("<br>", unsafe_allow_html=True)
    pres_cols = st.columns(2, gap="medium")
    with pres_cols[0]:
        st.markdown(
            """
<div class="presentation-card">
  <div class="eyebrow">Présentation</div>
  <ul>
    <li>Aggrégation d'actualités <b>Maroc · UE · monde</b>, organisées en 4 cadres de veille : <b>Réglementaire</b>, <b>Informative</b>, <b>Évènementielle</b>, <b>Concurrentielle</b>.</li>
    <li>Un <b>Signal du jour</b> : article à fort impact potentiel pour Les Domaines Agricoles.</li>
    <li><i>Thèmes couverts : Agrumes, Fruits rouges &amp; Tomates cerises · Produits laitiers &amp; Épicerie fine · Élevage (Ovins, Bovins, Caprins, Volailles &amp; Aquaculture) · Environnement, Eau &amp; Énergie · Normes : ESG, QSE &amp; SST.</i></li>
  </ul>
</div>
""",
            unsafe_allow_html=True,
        )
    with pres_cols[1]:
        st.markdown(
            """
<div class="presentation-card">
  <div class="eyebrow">Utiliser l'Agent</div>
  <ol>
    <li>Cliquer sur <b>« Veille DA »</b> pour revenir au tableau de bord</li>
    <li>Sélectionner la <b>période</b> souhaitée (max 30 jours)</li>
    <li>Choisir le <b>thème</b> à couvrir (un seul à la fois)</li>
    <li>Les actualités s'affichent <b>automatiquement</b> — cliquer sur <b>Voir tout</b> d'un cadre pour explorer les articles</li>
  </ol>
</div>
""",
            unsafe_allow_html=True,
        )

    df_urls = pd.DataFrame(MEDIA_SCOUT_SOURCE_CATALOG)
    # Retire les codes internes d'organisation (C1, A2, T4…) en tête de "Couverture"
    # — utiles côté maintenance mais déroutants pour l'utilisateur final.
    df_urls["Couverture"] = (
        df_urls["Couverture"].astype(str)
        .str.replace(r"^[A-Z]\d+\s*[—–\-:]?\s*", "", regex=True)
    )
    df_urls["Origine"] = df_urls["Journal"].map(get_source_origin)
    df_urls["Thème"] = df_urls["Journal"].map(MEDIA_SCOUT_FORCED_SOURCE_THEMES).map(_theme_display).fillna("Multi-thèmes")
    df_urls["Zone"] = df_urls["Journal"].map(MEDIA_SCOUT_SOURCE_ZONES).fillna("—")
    df_urls = df_urls.reindex(columns=["Journal", "Origine", "Thème", "Zone", "Couverture", "URL"])
    # Priorité d'affichage : MAROC > EU > WORLD, puis Thème, puis Journal
    _zone_order = {"MAROC": 0, "EU": 1, "WORLD": 2}
    df_urls["_zone_prio"] = df_urls["Zone"].map(_zone_order).fillna(99)
    df_urls = (
        df_urls.sort_values(["_zone_prio", "Thème", "Journal"], kind="mergesort")
        .drop(columns=["_zone_prio"])
        .reset_index(drop=True)
    )

    # Heading + download button (right-aligned)
    h_col1, h_col2 = st.columns([4, 1], vertical_alignment="bottom")
    with h_col1:
        st.markdown(
            f"""
<div class="sources-heading">
  Sources de référence
  <small>{len(MEDIA_SCOUT_SOURCE_CATALOG)} sources incluses · Maroc · UE · International</small>
</div>
""",
            unsafe_allow_html=True,
        )
    with h_col2:
        # Generate Excel : freeze first row + autofilter + auto-fit column widths + header style
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Excel export : on retire la colonne Zone (gardee uniquement pour le tri/in-app)
        df_excel = df_urls.drop(columns=["Zone"], errors="ignore")

        _buf = BytesIO()
        with pd.ExcelWriter(_buf, engine="openpyxl") as writer:
            df_excel.to_excel(writer, sheet_name="Sources", index=False)
            ws = writer.sheets["Sources"]

            # Freeze first row (header reste visible au scroll)
            ws.freeze_panes = "A2"

            # Auto-filter sur l'ensemble du tableau
            ws.auto_filter.ref = ws.dimensions

            # Auto-fit largeurs de colonnes (max contenu + marge, plafonne a 80)
            for col_idx, col_name in enumerate(df_excel.columns, start=1):
                if df_excel[col_name].empty:
                    max_len = len(str(col_name))
                else:
                    max_len = max(
                        df_excel[col_name].astype(str).map(len).max(),
                        len(str(col_name)),
                    )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 80)

            # Styling header row : bold + fond papier + bordure
            header_font = Font(bold=True, color="1A120A", size=11)
            header_fill = PatternFill(start_color="F7F2E6", end_color="F7F2E6", fill_type="solid")
            thin = Side(border_style="thin", color="C5A96D")
            header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = header_border

            # Hauteur de la 1ere ligne (header)
            ws.row_dimensions[1].height = 22

        st.download_button(
            label="⬇ Télécharger (Excel)",
            data=_buf.getvalue(),
            file_name=f"veille_ia_sources_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_sources_xlsx",
            width="stretch",
        )

    st.dataframe(
        df_urls.style.format(na_rep="No Data", precision=0),
        column_config={
            "Origine": st.column_config.TextColumn("Origine", width="small"),
            "Thème": st.column_config.TextColumn("Thème", width="medium"),
            "Zone": st.column_config.TextColumn("Zone", width="small"),
            "Couverture": st.column_config.TextColumn("Couverture", width="large"),
            "URL": st.column_config.LinkColumn(
                "URL",
                validate=r"^https?://.*$",
                max_chars=100,
                display_text="Ouvrir",
            ),
        },
        hide_index=True,
        width="stretch",
    )


# ─── Bottom spacing (2 empty rows at end of page) ─────────────────────────────
st.markdown('<div style="height:80px"></div>', unsafe_allow_html=True)


